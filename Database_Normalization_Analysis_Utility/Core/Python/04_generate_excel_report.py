"""
Script to generate an Excel report from functional dependency analysis results.
Creates a workbook with three sheets: Summary, Functional Dependencies, and Violations.
"""

import json
import sys
import re
from datetime import datetime

# Try to import openpyxl with helpful error message if not installed
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText
except ImportError:
    print("Error: openpyxl library is not installed.")
    print("Please install it using: pip install openpyxl")
    sys.exit(1)

from config_loader import ConfigLoader


def load_results(results_path):
    """
    Load the functional dependency analysis results from JSON file.

    Args:
        results_path: Path to the JSON results file

    Returns:
        Dictionary containing the analysis results

    Raises:
        FileNotFoundError: If results file doesn't exist
        json.JSONDecodeError: If file contains invalid JSON
    """
    try:
        with open(results_path, 'r') as f:
            results = json.load(f)
        print(f"Results loaded from: {results_path}")
        return results
    except FileNotFoundError:
        print(f"Error: Results file not found at {results_path}")
        print("Please run the 'Classify Dependencies' script first.")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON in results file: {e}")
        sys.exit(1)


def validate_excel_config(excel_config):
    """
    Validate that excel_config has all required fields.

    Args:
        excel_config: Dictionary containing Excel formatting configuration

    Raises:
        ValueError: If required fields are missing
    """
    required_fields = ['header_font_color', 'header_font_size', 'header_color',
                      'min_column_width', 'max_column_width', 'relevance_colors']

    for field in required_fields:
        if field not in excel_config:
            raise ValueError(f"Excel configuration missing required field: '{field}'")


def create_header_style(excel_config):
    """Create styling for header rows."""
    return {
        'font': Font(bold=True, color=excel_config['header_font_color'], size=excel_config['header_font_size']),
        'fill': PatternFill(start_color=excel_config['header_color'], end_color=excel_config['header_color'], fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_cell_style():
    """Create styling for data cells."""
    return {
        'alignment': Alignment(horizontal="left", vertical="center"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def validate_results_structure(results):
    """
    Validate that the results dictionary has the expected structure.

    Args:
        results: Dictionary loaded from JSON file

    Raises:
        ValueError: If required fields are missing or invalid
    """
    if not isinstance(results, dict):
        raise ValueError("Results must be a dictionary")

    # Check for required fields
    required_fields = ['functional_dependencies', 'table', 'database']
    for field in required_fields:
        if field not in results:
            raise ValueError(f"Results missing required field: '{field}'")

    fd_list = results.get('functional_dependencies')
    if not isinstance(fd_list, list):
        raise ValueError("'functional_dependencies' must be a list")

    if len(fd_list) == 0:
        raise ValueError("No functional dependencies found. Please run the analysis and classification scripts first.")


def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def auto_adjust_column_width(worksheet, excel_config):
    """Auto-adjust column widths based on content."""
    min_width = excel_config['min_column_width']
    max_width = excel_config['max_column_width']

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    # Calculate length, accounting for line breaks
                    cell_value = str(cell.value)
                    if '\n' in cell_value:
                        # For multi-line cells, use the longest line
                        max_line_length = max(len(line) for line in cell_value.split('\n'))
                        max_length = max(max_length, max_line_length)
                    else:
                        max_length = max(max_length, len(cell_value))
            except Exception:
                # Skip cells that can't be processed
                pass

        # Add padding and set reasonable min/max bounds from config
        adjusted_width = max(min_width, min(max_length + 3, max_width))
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_summary_sheet(workbook, results, transitive_count, composite_transitive_count, excel_config):
    """Create the summary sheet with overview information."""
    ws = workbook.active
    ws.title = "Summary"

    header_style = create_header_style(excel_config)
    cell_style = create_cell_style()

    # Title
    ws['A1'] = "Functional Dependency Analysis Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')

    # Format primary key for display
    primary_key = results.get('primarykey', 'N/A')
    if isinstance(primary_key, list):
        primary_key_display = ', '.join(primary_key)
    elif isinstance(primary_key, str):
        primary_key_display = primary_key
    else:
        primary_key_display = 'N/A'

    # Format unique keys for display
    unique_keys = results.get('uniquekey', [])
    if isinstance(unique_keys, list) and len(unique_keys) > 0:
        unique_key_parts = []
        for key in unique_keys:
            if isinstance(key, list):
                unique_key_parts.append('{' + ', '.join(key) + '}')
            elif isinstance(key, str) and key != "":
                unique_key_parts.append('{' + key + '}')
        unique_key_display = ', '.join(unique_key_parts) if unique_key_parts else 'None'
    else:
        unique_key_display = 'None'

    # Summary information
    row = 3
    summary_data = [
        ("Analysis Timestamp", results.get('analysis_timestamp', 'N/A')),
        ("Database", results.get('database', 'N/A')),
        ("Table", results.get('table', 'N/A')),
        ("Primary Key", primary_key_display),
        ("Unique Keys", unique_key_display),
        ("Total Combinations Checked", results.get('total_combinations_checked', 0)),
        ("Single Column Dependencies", len(results.get('single_column_dependencies', []))),
        ("Composite Dependencies", len(results.get('composite_dependencies', []))),
        ("Total Functional Dependencies", len(results.get('functional_dependencies', []))),
        ("Violations Found (No FD)", len(results.get('violations', []))),
        ("Transitive Dependencies Found", transitive_count),
        ("Composite Transitive Dependencies", composite_transitive_count),
        ("Errors", len(results.get('errors', []))),
    ]

    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    # Columns analyzed
    row += 1
    ws[f'A{row}'] = "Columns Analyzed"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for col in results.get('columns_analyzed', []):
        ws[f'A{row}'] = col
        row += 1

    # Relevance summary
    relevance_summary = results.get('relevance_summary', {})
    if relevance_summary:
        row += 1
        ws[f'A{row}'] = "Relevance Classification"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        for relevance_type, count in sorted(relevance_summary.items()):
            ws[f'A{row}'] = f"  {relevance_type}"
            ws[f'B{row}'] = count
            row += 1

    auto_adjust_column_width(ws, excel_config)


def format_determinant(determinant):
    """Format determinant for display (handle both single and composite)."""
    if isinstance(determinant, list):
        return ', '.join(determinant)
    return str(determinant)


def create_rich_text_for_column_list(text, primary_key):
    """
    Create rich text for comma-separated column names (without brackets/parentheses).
    Highlights individual PK columns in red only if ALL PK columns are present.

    Args:
        text: Comma-separated column names (e.g., "Manufacturer, Model, Color")
        primary_key: The primary key (string or list)

    Returns:
        CellRichText object with PK columns in red, or plain text
    """
    # Normalize primary key to list
    if isinstance(primary_key, str):
        pk_list = [primary_key] if primary_key else []
    else:
        pk_list = primary_key if primary_key else []

    if not pk_list or not text:
        return text

    # Split the text by comma to get individual columns
    columns = [col.strip() for col in text.split(',')]

    # Check if ALL PK columns are present
    pk_columns = [col for col in columns if col in pk_list]
    all_pk_present = set(pk_list).issubset(set(pk_columns))

    if not all_pk_present:
        # Not all PK columns present - return plain text
        return text

    # Build rich text with PK columns highlighted
    rich_text_parts = []
    for i, col in enumerate(columns):
        if i > 0:
            # Add comma and space between columns
            rich_text_parts.append(TextBlock(InlineFont(), ', '))

        # Highlight PK columns in red, others normal
        if col in pk_list:
            rich_text_parts.append(TextBlock(InlineFont(color='FF0000'), col))
        else:
            rich_text_parts.append(TextBlock(InlineFont(), col))

    return CellRichText(*rich_text_parts)


def create_rich_text_with_pk_highlight(text, primary_key):
    """
    Create rich text with primary key columns highlighted in red.

    Logic:
    - If brackets/parentheses contain ONLY PK columns: highlight entire bracketed section including brackets
    - If brackets/parentheses contain PK + other columns: highlight only individual PK column names

    Args:
        text: The text string to format
        primary_key: The primary key (string or list)

    Returns:
        CellRichText object with primary key portions in red
    """
    # Normalize primary key to list
    if isinstance(primary_key, str):
        pk_list = [primary_key] if primary_key else []
    else:
        pk_list = primary_key if primary_key else []

    if not pk_list:
        return text

    # Create a pattern to match both brackets [] and parentheses ()
    # Match patterns like [Manufacturer, Model] or (Manufacturer, Model)
    rich_text_parts = []
    last_end = 0

    # Find all bracketed or parenthesized sections
    for match in re.finditer(r'[\[\(]([^\]\)]+)[\]\)]', text):
        # Add text before the match (normal formatting)
        if match.start() > last_end:
            rich_text_parts.append(TextBlock(InlineFont(), text[last_end:match.start()]))

        # Get the opening and closing characters
        opening_char = match.group(0)[0]  # [ or (
        closing_char = match.group(0)[-1]  # ] or )

        # Check if this section contains primary key columns
        content = match.group(1)
        columns = [col.strip() for col in content.split(',')]

        # Separate PK columns from non-PK columns
        pk_columns = [col for col in columns if col in pk_list]
        non_pk_columns = [col for col in columns if col not in pk_list]

        # Check if ALL columns are PK columns and match the PK exactly
        is_exact_pk_match = (len(non_pk_columns) == 0 and
                            set(pk_columns) == set(pk_list) and
                            len(pk_columns) == len(pk_list))

        # Check if ALL PK columns are present (even if there are extra non-PK columns)
        all_pk_present = set(pk_list).issubset(set(pk_columns))

        if is_exact_pk_match:
            # Highlight entire bracketed/parenthesized section in red (including brackets/parentheses)
            rich_text_parts.append(TextBlock(InlineFont(color='FF0000'), match.group(0)))
        elif all_pk_present:
            # Mixed: ALL PK columns present + other columns
            # Highlight opening bracket/parenthesis normally
            rich_text_parts.append(TextBlock(InlineFont(), opening_char))

            # Process each column individually
            for i, col in enumerate(columns):
                if i > 0:
                    # Add comma and space between columns
                    rich_text_parts.append(TextBlock(InlineFont(), ', '))

                # Highlight PK columns in red, others normal
                if col in pk_list:
                    rich_text_parts.append(TextBlock(InlineFont(color='FF0000'), col))
                else:
                    rich_text_parts.append(TextBlock(InlineFont(), col))

            # Add closing bracket/parenthesis normally
            rich_text_parts.append(TextBlock(InlineFont(), closing_char))
        else:
            # Not all PK columns present - normal formatting for entire section
            rich_text_parts.append(TextBlock(InlineFont(), match.group(0)))

        last_end = match.end()

    # Add remaining text after last match
    if last_end < len(text):
        rich_text_parts.append(TextBlock(InlineFont(), text[last_end:]))

    # If we found any matches, return rich text, otherwise return plain text
    if rich_text_parts:
        return CellRichText(*rich_text_parts)
    else:
        return text


def create_functional_dependencies_sheet(workbook, results, excel_config):
    """Create sheet with functional dependencies found."""
    ws = workbook.create_sheet("Functional Dependencies")

    header_style = create_header_style(excel_config)
    cell_style = create_cell_style()

    # Get primary key for highlighting
    primary_key = results.get('primarykey', '')

    # Headers
    headers = ["Determinant (A)", "Type", "Dependent (B)", "Relevance", "Relevance Reason", "Description", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_style(cell, header_style)

    # Data
    fd_list = results.get('functional_dependencies', [])
    for row_num, fd in enumerate(fd_list, 2):
        determinant = fd.get('determinant', '')
        det_size = fd.get('determinant_size', 1)
        relevance = fd.get('relevance', 'UNKNOWN')

        # Apply rich text formatting to Determinant (A) - column 1
        determinant_text = format_determinant(determinant)
        ws.cell(row=row_num, column=1, value=create_rich_text_for_column_list(determinant_text, primary_key))

        ws.cell(row=row_num, column=2, value="Composite" if det_size > 1 else "Single")
        ws.cell(row=row_num, column=3, value=fd.get('dependent', ''))
        ws.cell(row=row_num, column=4, value=relevance)

        # Apply rich text formatting to Relevance Reason - column 5
        relevance_reason_text = fd.get('relevance_reason', '')
        ws.cell(row=row_num, column=5, value=create_rich_text_with_pk_highlight(relevance_reason_text, primary_key))

        # Apply rich text formatting to description (highlight PK in red)
        description_text = fd.get('description', '')
        ws.cell(row=row_num, column=6, value=create_rich_text_with_pk_highlight(description_text, primary_key))

        ws.cell(row=row_num, column=7, value=fd.get('status', ''))

        # Color code by relevance using config colors
        relevance_cell = ws.cell(row=row_num, column=4)
        relevance_colors = excel_config['relevance_colors']
        if relevance in relevance_colors:
            color = relevance_colors[relevance]
            relevance_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for col_num in range(1, 8):
            apply_style(ws.cell(row=row_num, column=col_num), cell_style)

    # Add AutoFilter to header row
    ws.auto_filter.ref = ws.dimensions

    auto_adjust_column_width(ws, excel_config)


def create_violations_sheet(workbook, results, excel_config):
    """Create sheet with violations (no functional dependencies)."""
    ws = workbook.create_sheet("No Functional Dependencies")

    header_style = create_header_style(excel_config)
    cell_style = create_cell_style()

    # Headers
    headers = ["Determinant (A)", "Type", "Dependent (B)", "Description", "Status"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_style(cell, header_style)

    # Data
    violations_list = results.get('violations', [])
    for row_num, violation in enumerate(violations_list, 2):
        determinant = violation.get('determinant', '')
        det_size = violation.get('determinant_size', 1)

        ws.cell(row=row_num, column=1, value=format_determinant(determinant))
        ws.cell(row=row_num, column=2, value="Composite" if det_size > 1 else "Single")
        ws.cell(row=row_num, column=3, value=violation.get('dependent', ''))
        ws.cell(row=row_num, column=4, value=violation.get('description', ''))
        ws.cell(row=row_num, column=5, value=violation.get('status', ''))

        for col_num in range(1, 6):
            apply_style(ws.cell(row=row_num, column=col_num), cell_style)

    # Add AutoFilter to header row
    ws.auto_filter.ref = ws.dimensions

    auto_adjust_column_width(ws, excel_config)


def find_transitive_chain(start, current, fd_map, pk_list, visited, max_depth=10):
    """
    Recursively find all transitive dependency chains starting from a column.
    Returns list of chains where each chain is a list of columns.
    """
    if max_depth <= 0:
        return []

    chains = []

    # Get all columns that current determines
    dependents = fd_map.get(current, [])

    for dependent in dependents:
        # Skip if dependent is in primary key
        if dependent in pk_list:
            continue

        # Skip if we've already visited (avoid cycles)
        if dependent in visited:
            continue

        # Skip if dependent is the start (circular)
        if dependent == start:
            continue

        # Create new chain with this dependent
        new_visited = visited | {dependent}

        # Recursively find chains from this dependent
        sub_chains = find_transitive_chain(start, dependent, fd_map, pk_list, new_visited, max_depth - 1)

        if sub_chains:
            # Add current dependent to the beginning of each sub-chain
            for sub_chain in sub_chains:
                chains.append([dependent] + sub_chain)
        else:
            # This is a leaf node, create a chain with just this dependent
            chains.append([dependent])

    return chains


def find_transitive_dependencies(results):
    """
    Find transitive dependencies from functional dependencies.
    Detects chains of any length: A -> B -> C, A -> B -> C -> D, etc.

    A transitive dependency exists when:
    - A -> B -> ... -> Z (chain of dependencies)
    - All intermediate columns are non-key attributes
    - A also directly determines Z

    This indicates: A -> B -> ... -> Z (transitive dependency)
    """
    fd_list = results.get('functional_dependencies', [])
    primary_key = results.get('primarykey', '')

    # Normalize primary key to list for easier checking
    if isinstance(primary_key, str):
        pk_list = [primary_key] if primary_key else []
    else:
        pk_list = primary_key if primary_key else []

    # Build a dictionary of functional dependencies for quick lookup
    # Only include single-column determinants for transitive chains
    fd_map = {}
    for fd in fd_list:
        determinant = fd.get('determinant')
        dependent = fd.get('dependent')

        # Only consider single-column determinants for classic transitive dependencies
        if isinstance(determinant, list):
            continue

        if determinant not in fd_map:
            fd_map[determinant] = []
        fd_map[determinant].append(dependent)

    transitive_deps = []
    seen_chains = set()  # To avoid duplicate chains

    # For each single-column determinant
    for fd in fd_list:
        a = fd.get('determinant')  # A (starting column)

        # Only consider single-column A
        if isinstance(a, list):
            continue

        # Skip if A is not in fd_map (doesn't determine anything)
        if a not in fd_map:
            continue

        # Find all transitive chains starting from A
        chains = find_transitive_chain(a, a, fd_map, pk_list, {a}, max_depth=10)

        # Process each chain
        for chain in chains:
            # Chain must be at least length 2 (B -> C minimum)
            if len(chain) < 2:
                continue

            # Check if A directly determines the final column in the chain
            final_col = chain[-1]
            if final_col not in fd_map.get(a, []):
                continue

            # Build the full chain: A -> B -> C -> ... -> Z
            full_chain = [a] + chain
            # Add brackets around each element in the chain
            chain_str = ' -> '.join([f'[{col}]' for col in full_chain])

            # Avoid duplicates
            if chain_str in seen_chains:
                continue
            seen_chains.add(chain_str)

            # Extract intermediate columns (all except first and last)
            intermediate_cols = full_chain[1:-1]

            transitive_deps.append({
                'a': a,
                'intermediates': intermediate_cols,
                'final': final_col,
                'chain': chain_str,
                'degree': len(full_chain),
                'description': f"Transitive (degree {len(full_chain)}): [{a}] determines [{final_col}] through {' -> '.join([f'[{col}]' for col in intermediate_cols])}"
            })

    return transitive_deps


def create_transitive_dependencies_sheet(workbook, results, excel_config):
    """Create sheet with transitive dependencies (both single and composite)."""
    ws = workbook.create_sheet("Transitive Dependencies")

    header_style = create_header_style(excel_config)
    cell_style = create_cell_style()

    # Get primary key for highlighting
    primary_key = results.get('primarykey', '')

    # Headers
    headers = ["Start Column", "Intermediate Columns", "Final Column", "Degree", "Dependency Chain", "Description", "Type"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        apply_style(cell, header_style)

    # Find both single-column and composite transitive dependencies
    transitive_deps = find_transitive_dependencies(results)
    composite_transitive_deps = find_composite_transitive_dependencies(results)

    # Combine both types
    all_transitive_deps = []

    # Add single-column transitive deps
    for td in transitive_deps:
        all_transitive_deps.append({
            'start': td.get('a', ''),
            'intermediates': ' -> '.join(td.get('intermediates', [])),
            'final': td.get('final', ''),
            'degree': td.get('degree', ''),
            'chain': td.get('chain', ''),
            'description': td.get('description', ''),
            'type': 'Single-Column'
        })

    # Add composite transitive deps
    for ctd in composite_transitive_deps:
        dep_type = 'Composite (2NF Violation)' if ctd.get('is_partial_dependency', False) else 'Composite'
        all_transitive_deps.append({
            'start': ctd.get('composite', ''),
            'intermediates': ctd.get('intermediate', ''),
            'final': ctd.get('final', ''),
            'degree': '2',  # Composite -> Intermediate -> Final
            'chain': ctd.get('chain', ''),
            'description': ctd.get('description', ''),
            'type': dep_type,
            'is_partial': ctd.get('is_partial_dependency', False)
        })

    # Data
    for row_num, td in enumerate(all_transitive_deps, 2):
        # Apply rich text formatting to Start Column (A) - column 1
        start_text = td.get('start', '')
        ws.cell(row=row_num, column=1, value=create_rich_text_for_column_list(start_text, primary_key))

        ws.cell(row=row_num, column=2, value=td.get('intermediates', ''))
        ws.cell(row=row_num, column=3, value=td.get('final', ''))
        ws.cell(row=row_num, column=4, value=td.get('degree', ''))

        # Apply rich text formatting to chain and description (highlight PK in red)
        chain_text = td.get('chain', '')
        description_text = td.get('description', '')

        ws.cell(row=row_num, column=5, value=create_rich_text_with_pk_highlight(chain_text, primary_key))
        ws.cell(row=row_num, column=6, value=create_rich_text_with_pk_highlight(description_text, primary_key))
        ws.cell(row=row_num, column=7, value=td.get('type', ''))

        for col_num in range(1, 8):
            apply_style(ws.cell(row=row_num, column=col_num), cell_style)

    # Add note if no transitive dependencies found
    if not all_transitive_deps:
        ws.cell(row=2, column=1, value="No transitive dependencies found")
        ws.merge_cells('A2:G2')
        ws.cell(row=2, column=1).font = Font(italic=True)
    else:
        # Add AutoFilter to header row
        ws.auto_filter.ref = ws.dimensions

    auto_adjust_column_width(ws, excel_config)


def find_composite_transitive_dependencies(results):
    """
    Find transitive dependencies involving composite keys.
    Patterns detected:
    1. (A, B) -> C -> D (composite determines C, C determines D, composite determines D)
    2. A -> (B, C) -> D (A determines composite, composite determines D, A determines D)
    3. (A, B) -> (C, D) -> E (composite to composite to single)

    For simplicity, we focus on pattern 1: composite -> single -> single
    """
    fd_list = results.get('functional_dependencies', [])
    primary_key = results.get('primarykey', '')

    # Normalize primary key to list
    if isinstance(primary_key, str):
        pk_list = [primary_key] if primary_key else []
    else:
        pk_list = primary_key if primary_key else []

    # Build fd_map with both single and composite determinants
    fd_map = {}
    for fd in fd_list:
        determinant = fd.get('determinant')
        dependent = fd.get('dependent')

        # Create hashable key
        if isinstance(determinant, list):
            det_key = tuple(determinant)
        else:
            det_key = determinant

        if det_key not in fd_map:
            fd_map[det_key] = []
        fd_map[det_key].append(dependent)

    composite_transitive_deps = []

    # Pattern 1: (A, B) -> C -> D
    # This includes cases where C is part of the composite key (important for 2NF violations)
    for fd in fd_list:
        determinant = fd.get('determinant')
        c = fd.get('dependent')  # C (intermediate column)

        # Only consider composite determinants
        if not isinstance(determinant, list):
            continue

        # Check if C (as single column) determines any D
        if c in fd_map:
            for d in fd_map[c]:
                # Skip if D is same as C
                if d == c:
                    continue

                # Skip if D is same as any column in determinant
                if isinstance(determinant, list) and d in determinant:
                    continue

                # Check if composite also determines D (completing transitive chain)
                det_tuple = tuple(determinant) if isinstance(determinant, list) else determinant
                if d in fd_map.get(det_tuple, []):
                    # Found: (A, B) -> C -> D
                    # Determine if this is a partial dependency (2NF violation)
                    is_partial = c in pk_list
                    dep_type = "Partial Dependency (2NF Violation)" if is_partial else "Composite Transitive"

                    composite_transitive_deps.append({
                        'composite': format_determinant(determinant),
                        'intermediate': c,
                        'final': d,
                        'chain': f"[{format_determinant(determinant)}] -> [{c}] -> [{d}]",
                        'description': f"{dep_type}: [{format_determinant(determinant)}] determines [{d}] through [{c}]",
                        'is_partial_dependency': is_partial
                    })

    return composite_transitive_deps


def main():
    """Main function to generate Excel report."""
    print("=" * 80)
    print("Excel Report Generator")
    print("=" * 80)

    try:
        # Load configuration using ConfigLoader
        config_loader = ConfigLoader()
        results_path = config_loader.get_functional_dependencies_path()
        excel_config = config_loader.get_excel_config()

        # Validate that results file path exists
        from pathlib import Path
        if not Path(results_path).exists():
            print(f"Error: Results file not found at {results_path}")
            print("Please run the 'Classify Dependencies' script first.")
            sys.exit(1)

        # Validate Excel configuration
        validate_excel_config(excel_config)

        # Load results
        results = load_results(results_path)

        # Validate results structure
        validate_results_structure(results)

        # Create output filename with table name and timestamp
        table_name = results.get('table', 'unknown')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"{table_name}_functional_dependencies_{timestamp}.xlsx"
        output_path = config_loader.get_output_path(output_filename)

        # Create output directory if it doesn't exist
        output_dir = Path(output_path).parent
        output_dir.mkdir(parents=True, exist_ok=True)

        # Create workbook
        print(f"\nCreating Excel workbook...")
        workbook = openpyxl.Workbook()

        # Find transitive dependencies first (needed for summary)
        print("Analyzing transitive dependencies...")
        transitive_deps = find_transitive_dependencies(results)
        transitive_count = len(transitive_deps)

        # Find composite transitive dependencies (needed for summary)
        composite_transitive_deps = find_composite_transitive_dependencies(results)
        composite_transitive_count = len(composite_transitive_deps)

        # Create sheets
        print("Creating Summary sheet...")
        create_summary_sheet(workbook, results, transitive_count, composite_transitive_count, excel_config)

        print("Creating Functional Dependencies sheet...")
        create_functional_dependencies_sheet(workbook, results, excel_config)

        print("Creating Transitive Dependencies sheet...")
        create_transitive_dependencies_sheet(workbook, results, excel_config)

        print("Creating No Functional Dependencies sheet...")
        create_violations_sheet(workbook, results, excel_config)

        # Save workbook
        print(f"\nSaving Excel workbook...")
        workbook.save(output_path)

        print("=" * 80)
        print("Report generation completed successfully!")
        print("=" * 80)
        print(f"\nExcel report saved to: {output_path}")
        print("Use the UI to open the report.")

    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)
    except PermissionError:
        print(f"Error: Permission denied writing to {output_path}")
        print("Please close the Excel file if it's open and try again.")
        sys.exit(1)
    except IOError as e:
        print(f"Error: Cannot write Excel file: {e}")
        sys.exit(1)
    except Exception as e:
        print("=" * 80)
        print(f"ERROR: Report generation failed: {e}")
        print("=" * 80)
        sys.exit(1)


if __name__ == "__main__":
    main()
