"""
Script to format the Final Excel Report based on part naming convention.

This script reads the part_naming_convention from config.json and modifies
the Database_Object and object_name columns in the Forward_Dependencies and
Reverse_Dependencies tabs to show only the relevant parts of the object name.

Naming convention options:
- 1: Show only object name (from database.schema.object_name)
- 2: Show schema.object_name (from database.schema.object_name)
- 3: Show database.schema.object_name (no change)
"""

import os
import sys
import logging
from datetime import datetime
from config_loader import ConfigLoader

def parse_object_name(full_name, convention):
    """
    Parse object name based on naming convention.

    Args:
        full_name: Full object name in format database.schema.object_name
        convention: 1, 2, or 3

    Returns:
        Formatted object name based on convention
    """
    if not full_name or full_name == '':
        return full_name

    # Split by period
    parts = str(full_name).split('.')

    if convention == 1:
        # Return only object name (last part)
        return parts[-1] if parts else full_name
    elif convention == 2:
        # Return schema.object_name (last 2 parts)
        if len(parts) >= 2:
            return '.'.join(parts[-2:])
        else:
            return full_name
    elif convention == 3:
        # Return full name (database.schema.object_name)
        return full_name
    else:
        # Invalid convention, return full name
        return full_name

def remove_object_descriptions_from_path(path_value, is_reverse=False, convention=3, remove_descriptions=True):
    """
    Remove object type descriptions from object_name_path and apply naming convention.

    Args:
        path_value: String containing object path with arrows (➡️ or ⬅️)
        is_reverse: Boolean indicating if this is a reverse dependency (uses ⬅️)
        convention: Naming convention (1, 2, or 3)
        remove_descriptions: Whether to remove object type descriptions

    Returns:
        Formatted path with descriptions removed (except UNKNOWN) and naming convention applied

    Example (Forward, convention=1, remove_descriptions=True):
        Input:  "QA07_Greg.dbo.spAR_AncillaryCapDetail_Trigger_Insert.SQL_STORED_PROCEDURE ➡️ QA07_Greg.dbo.AR_AncillaryCapDetail.USER_TABLE"
        Output: "spAR_AncillaryCapDetail_Trigger_Insert ➡️ AR_AncillaryCapDetail"

        Input:  "QA07_Greg.dbo.spAR_AncillaryPayer_Inherit_Generic_Staged.SQL_STORED_PROCEDURE ➡️ s.UNKNOWN"
        Output: "spAR_AncillaryPayer_Inherit_Generic_Staged ➡️ s.UNKNOWN"

    Example (Reverse, convention=2, remove_descriptions=True):
        Input:  "QA07_Greg.dbo.spAR_AncillaryCapDetail_Delete.SQL_STORED_PROCEDURE ⬅️ QA07_Greg.dbo.spAR_AncillaryCapDetail_Trigger_Insert"
        Output: "dbo.spAR_AncillaryCapDetail_Delete ⬅️ dbo.spAR_AncillaryCapDetail_Trigger_Insert"
    """
    if not path_value or path_value == '':
        return path_value

    # Determine which arrow to use
    arrow = '⬅️' if is_reverse else '➡️'

    # Split by arrow
    parts = str(path_value).split(arrow)

    formatted_parts = []
    for part in parts:
        part = part.strip()

        # Check if this part ends with UNKNOWN
        if part.endswith('.UNKNOWN') or part == 'UNKNOWN':
            # Keep UNKNOWN as-is
            formatted_parts.append(part)
        else:
            # First, remove the object type description if enabled
            if remove_descriptions:
                # Split by period
                segments = part.split('.')
                if len(segments) > 1:
                    # Check if the last segment looks like an object type description
                    # (all caps with underscores, like SQL_STORED_PROCEDURE, USER_TABLE, etc.)
                    last_segment = segments[-1]
                    if last_segment.isupper() and '_' in last_segment:
                        # Remove the description
                        part = '.'.join(segments[:-1])

            # Now apply the naming convention
            part = parse_object_name(part, convention)
            formatted_parts.append(part)

    return f' {arrow} '.join(formatted_parts)

def remove_object_description_from_fullname(fullname_value):
    """
    Remove object type description from referenced_object_fullname or referencing_object_fullname.

    Args:
        fullname_value: String containing object fullname with description

    Returns:
        Formatted fullname with description removed (except UNKNOWN)

    Example:
        Input:  "QA07_Greg.dbo.AR_AncillaryCapDetail.USER_TABLE"
        Output: "QA07_Greg.dbo.AR_AncillaryCapDetail"

        Input:  "s.UNKNOWN"
        Output: "s.UNKNOWN"
    """
    if not fullname_value or fullname_value == '':
        return fullname_value

    # Convert to string
    fullname_str = str(fullname_value).strip()

    # Check if this ends with UNKNOWN
    if fullname_str.endswith('.UNKNOWN') or fullname_str == 'UNKNOWN':
        # Keep UNKNOWN as-is
        return fullname_str

    # Split by period
    segments = fullname_str.split('.')

    if len(segments) > 1:
        # Check if the last segment looks like an object type description
        # (all caps with underscores, like SQL_STORED_PROCEDURE, USER_TABLE, etc.)
        last_segment = segments[-1]
        if last_segment.isupper() and ('_' in last_segment or last_segment in ['VIEW', 'TABLE', 'PROCEDURE', 'FUNCTION']):
            # Remove the description
            return '.'.join(segments[:-1])
        else:
            # Keep as-is
            return fullname_str
    else:
        # No periods, keep as-is
        return fullname_str

def find_column_indices(headers):
    """
    Find column indices for all columns that need formatting.

    Args:
        headers: List of header values from the worksheet

    Returns:
        Dictionary mapping column names to their indices (1-based)
    """
    column_map = {
        'Database_Object': None,
        'object_name': None,
        'object_name_path': None,
        'referenced_object_fullname': None,
        'referencing_object_fullname': None
    }

    for idx, header in enumerate(headers, start=1):
        if header in column_map:
            column_map[header] = idx

    return column_map

def format_column(worksheet, column_idx, column_name, max_row, formatter_func, log):
    """
    Format a single column in the worksheet.

    Args:
        worksheet: The openpyxl worksheet object
        column_idx: Column index (1-based)
        column_name: Name of the column (for logging)
        max_row: Maximum row number to process
        formatter_func: Function to apply to each cell value
        log: Logger object

    Returns:
        Number of rows formatted
    """
    if not column_idx:
        return 0

    log.info(f"  Formatting {column_name} column (column {column_idx})...")
    row_count = 0

    for row in range(2, max_row + 1):
        cell = worksheet.cell(row=row, column=column_idx)
        if cell.value:
            try:
                cell.value = formatter_func(cell.value)
                row_count += 1
            except Exception as e:
                log.warning(f"    Error formatting row {row}: {e}")

    log.info(f"    Formatted {row_count} rows")
    return row_count

def process_dependency_tab(worksheet, tab_name, is_reverse, convention, remove_descriptions, log):
    """
    Process a single dependency tab (Forward or Reverse).

    Args:
        worksheet: The openpyxl worksheet object
        tab_name: Name of the tab (for logging)
        is_reverse: Boolean indicating if this is a reverse dependency tab
        convention: Naming convention (1, 2, or 3)
        remove_descriptions: Whether to remove object type descriptions
        log: Logger object

    Returns:
        Total number of cells formatted
    """
    log.info(f"Processing {tab_name} tab...")

    # Get headers and find column indices
    headers = [cell.value for cell in worksheet[1]]
    columns = find_column_indices(headers)

    total_formatted = 0

    # Format Database_Object column
    if columns['Database_Object']:
        count = format_column(
            worksheet,
            columns['Database_Object'],
            'Database_Object',
            worksheet.max_row,
            lambda val: parse_object_name(val, convention),
            log
        )
        total_formatted += count

    # Format object_name column
    if columns['object_name']:
        count = format_column(
            worksheet,
            columns['object_name'],
            'object_name',
            worksheet.max_row,
            lambda val: parse_object_name(val, convention),
            log
        )
        total_formatted += count

    # Format object_name_path column
    if columns['object_name_path']:
        count = format_column(
            worksheet,
            columns['object_name_path'],
            'object_name_path',
            worksheet.max_row,
            lambda val: remove_object_descriptions_from_path(val, is_reverse, convention, remove_descriptions),
            log
        )
        total_formatted += count

    # Format referenced_object_fullname column
    if columns['referenced_object_fullname']:
        def format_referenced(val):
            val_str = str(val)
            if remove_descriptions:
                val_str = remove_object_description_from_fullname(val_str)
            return parse_object_name(val_str, convention)

        count = format_column(
            worksheet,
            columns['referenced_object_fullname'],
            'referenced_object_fullname',
            worksheet.max_row,
            format_referenced,
            log
        )
        total_formatted += count

    # Format referencing_object_fullname column
    if columns['referencing_object_fullname']:
        def format_referencing(val):
            val_str = str(val)
            if remove_descriptions:
                val_str = remove_object_description_from_fullname(val_str)
            return parse_object_name(val_str, convention)

        count = format_column(
            worksheet,
            columns['referencing_object_fullname'],
            'referencing_object_fullname',
            worksheet.max_row,
            format_referencing,
            log
        )
        total_formatted += count

    log.info(f"  Total cells formatted in {tab_name}: {total_formatted}")
    return total_formatted

def format_excel_file(input_file, output_file, convention, remove_descriptions, log):
    """Format Excel file based on naming convention and description removal setting."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl library not found. Please install it: pip install openpyxl")
        return False

    try:
        log.info(f"Loading Excel file: {input_file}")

        if not os.path.exists(input_file):
            log.error(f"Input file not found: {input_file}")
            return False

        # Load workbook
        wb = load_workbook(input_file)

        log.info(f"Using part naming convention: {convention}")
        log.info(f"Remove object descriptions: {remove_descriptions}")
        log.info("")

        # Process Forward_Dependencies tab
        if "Forward_Dependencies" in wb.sheetnames:
            ws = wb["Forward_Dependencies"]
            process_dependency_tab(ws, "Forward_Dependencies", is_reverse=False,
                                  convention=convention, remove_descriptions=remove_descriptions, log=log)
        else:
            log.warning("Forward_Dependencies tab not found in workbook")

        # Process Reverse_Dependencies tab
        if "Reverse_Dependencies" in wb.sheetnames:
            ws = wb["Reverse_Dependencies"]
            process_dependency_tab(ws, "Reverse_Dependencies", is_reverse=True,
                                  convention=convention, remove_descriptions=remove_descriptions, log=log)
        else:
            log.warning("Reverse_Dependencies tab not found in workbook")

        # Save formatted workbook
        log.info("")
        log.info(f"Saving formatted Excel file to: {output_file}")
        wb.save(output_file)
        log.info("Excel file formatted successfully!")

        return True

    except Exception as e:
        log.error(f"Error formatting Excel file: {str(e)}")
        import traceback
        log.error(traceback.format_exc())
        return False

def main():
    # Load configuration
    config = ConfigLoader()

    # Get configuration values
    output_dir = config.get_output_dir()
    log_dir = config.get_log_dir()

    # Get input and output file paths
    input_file = os.path.join(output_dir, config.get_final_excel_report())
    output_file = os.path.join(output_dir, config.get_final_excel_report_formatted())

    # Get naming convention
    try:
        convention = int(config.get_part_naming_convention())
        if convention not in [1, 2, 3]:
            print(f"ERROR: Invalid part_naming_convention value: {convention}")
            print("Valid values are 1, 2, or 3")
            sys.exit(1)
    except Exception as e:
        print(f"ERROR: Could not read part_naming_convention from config.json: {e}")
        sys.exit(1)

    # Get remove_object_description setting
    try:
        remove_descriptions_value = config.get_remove_object_description()
        # Handle both boolean and string values
        if isinstance(remove_descriptions_value, bool):
            remove_descriptions = remove_descriptions_value
        else:
            remove_descriptions = str(remove_descriptions_value).lower() in ('true', 'yes', '1')
    except Exception as e:
        print(f"ERROR: Could not read remove_object_description from config.json: {e}")
        sys.exit(1)

    # Setup logging
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    log_filename = f'log_07_format_excel_file_{timestamp}.log'
    log_filepath = os.path.join(log_dir, log_filename)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filepath, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    logging.info("="*60)
    logging.info("Formatting Excel Report")
    logging.info("="*60)
    logging.info(f"Input file:                  {input_file}")
    logging.info(f"Output file:                 {output_file}")
    logging.info(f"Naming convention:           {convention}")
    logging.info(f"Remove object descriptions:  {remove_descriptions}")

    if convention == 1:
        logging.info("  Format: object_name only")
    elif convention == 2:
        logging.info("  Format: schema.object_name")
    elif convention == 3:
        logging.info("  Format: database.schema.object_name (full)")

    logging.info("")

    # Format Excel file
    success = format_excel_file(input_file, output_file, convention, remove_descriptions, logging)

    if success:
        logging.info("")
        logging.info("="*60)
        logging.info("COMPLETE!")
        logging.info("="*60)
        logging.info(f"Formatted Excel file: {output_file}")
        logging.info(f"Log file:             {log_filepath}")
    else:
        logging.error("Failed to format Excel file")
        sys.exit(1)

if __name__ == "__main__":
    main()


