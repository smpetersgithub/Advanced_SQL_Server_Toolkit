"""
Script to scan Java source files and extract a complete UI-to-stored-procedure mapping.
Scans DAO files for stored procedure calls and UI files (Handlers, Actions, Controllers,
Services) for DAO usage, then builds and exports a full Stored Procedure -> DAO -> UI
Component mapping to CSV and Excel.
"""

import os
import re
import csv
import sys
from collections import defaultdict
import logging
from config_loader import ConfigLoader

logger = logging.getLogger(__name__)

# Patterns
sp_pattern1 = re.compile(r'prepareCall\s*\(\s*["\{]+\s*(?:\?=)?call\s+(?:dbo\.)?([a-zA-Z0-9_]+)', re.IGNORECASE)
sp_pattern2 = re.compile(r'getCallableStatement\s*\(\s*["\{]+\s*(?:\?=)?call\s+(?:dbo\.)?([a-zA-Z0-9_]+)', re.IGNORECASE)
sp_pattern3 = re.compile(r'execute\s*\(\s*["\']([a-zA-Z0-9_]+)["\']', re.IGNORECASE)
sp_pattern4 = re.compile(r'super\s*\(\s*ds\s*,\s*["\']([a-zA-Z0-9_]+)["\']', re.IGNORECASE)
sp_pattern5 = re.compile(r'SQL\s*=\s*["\'](?:dbo\.)?([a-zA-Z0-9_]+)["\']', re.IGNORECASE)

class_pattern = re.compile(r'public\s+class\s+([a-zA-Z0-9_]+)')
import_pattern = re.compile(r'import\s+[\w.]+\.([a-zA-Z0-9_]+DAO);')
new_dao_pattern = re.compile(r'new\s+([a-zA-Z0-9_]+DAO)\s*\(')

# Data structures
dao_to_stored_procs = defaultdict(set)  # DAO -> Set of stored procs
dao_to_file = {}  # DAO -> file path
ui_to_daos = defaultdict(set)  # UI component -> Set of DAOs used
ui_to_file = {}  # UI component -> file path
ui_types = {}  # UI component -> type (Handler, Action, etc.)

def scan_dao_file(filepath, dao_count_tracker):
    """Scan a DAO file for stored procedures"""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        class_match = class_pattern.search(content)
        if not class_match:
            return

        class_name = class_match.group(1)

        # Only process DAO classes (including DAOSql, DAOImpl, etc.)
        if 'DAO' not in class_name:
            return

        dao_count_tracker[0] += 1

        # Find all stored procedures
        sp_matches = []
        sp_matches.extend(sp_pattern1.findall(content))
        sp_matches.extend(sp_pattern2.findall(content))

        for match in sp_pattern3.findall(content):
            if match.lower().startswith('sp'):
                sp_matches.append(match)
        for match in sp_pattern4.findall(content):
            if match.lower().startswith('sp'):
                sp_matches.append(match)
        for match in sp_pattern5.findall(content):
            if match.lower().startswith('sp'):
                sp_matches.append(match)

        if sp_matches:
            dao_to_file[class_name] = filepath
            dao_to_stored_procs[class_name].update(sp_matches)

    except Exception as e:
        logger.warning(f"Failed to scan DAO file {filepath}: {e}")

def scan_ui_file(filepath):
    """Scan UI files (Handlers, Actions, etc.) for DAO usage"""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        class_match = class_pattern.search(content)
        if not class_match:
            return

        class_name = class_match.group(1)

        # Determine UI type
        ui_type = 'Unknown'
        if 'Handler' in class_name:
            ui_type = 'Page Handler'
        elif 'Action' in class_name:
            ui_type = 'Struts Action'
        elif 'Controller' in class_name:
            ui_type = 'Spring Controller'
        elif 'Service' in class_name or 'ServiceImpl' in class_name:
            ui_type = 'Service'
        else:
            return  # Skip non-UI files

        # Find DAO imports
        imported_daos = set(import_pattern.findall(content))

        # Find DAO instantiations (new SomeDAO())
        instantiated_daos = set(new_dao_pattern.findall(content))

        # Combine both
        used_daos = imported_daos | instantiated_daos

        if used_daos:
            ui_to_file[class_name] = filepath
            ui_types[class_name] = ui_type
            ui_to_daos[class_name].update(used_daos)

    except Exception as e:
        logger.warning(f"Failed to scan UI file {filepath}: {e}")

def main():
    # Load configuration
    try:
        config = ConfigLoader()
    except (FileNotFoundError, ValueError, KeyError) as e:
        print(f"[ERROR] Configuration error: {e}")
        sys.exit(1)

    # Get configuration values
    java_source_dirs = config.get('paths', 'java_source_dirs', [])
    output_dir = config.get_output_dir()
    csv_filename = config.get('files', 'complete_mapping_csv')
    excel_filename = config.get('files', 'complete_mapping_excel')

    # Setup logging
    log_file = config.setup_logging('01_extract_complete_ui_mapping')

    logger.info("=" * 80)
    logger.info("Database Object Dependency Utility - Complete UI Mapping Generator")
    logger.info("Stored Procedure -> DAO -> UI Handler/Action -> JSP")
    logger.info("=" * 80)

    # Validate configuration
    if not java_source_dirs:
        logger.warning("No Java source directories configured in config.json")
        logger.warning("Please add 'java_source_dirs' to the 'paths' section")
        logger.info("\nExiting - no source directories to scan")
        return

    logger.info("\n[1/4] Scanning DAO files for stored procedures...")
    dao_count_tracker = [0]

    # Scan DAO files (including DAOSql.java, DAOImpl.java, etc.)
    for base_dir in java_source_dirs:
        if os.path.exists(base_dir):
            for root, dirs, files in os.walk(base_dir):
                for file in files:
                    if 'DAO' in file and file.endswith('.java'):
                        scan_dao_file(os.path.join(root, file), dao_count_tracker)

    dao_count = dao_count_tracker[0]
    logger.info(f"   Found {dao_count} DAO files")
    logger.info(f"   Found {len(dao_to_stored_procs)} DAOs with stored procedures")

    logger.info("\n[2/4] Scanning UI files (Handlers, Actions, Services) for DAO usage...")

    for base_dir in java_source_dirs:
        if os.path.exists(base_dir):
            for root, dirs, files in os.walk(base_dir):
                for file in files:
                    if file.endswith('.java') and not file.endswith('DAO.java'):
                        scan_ui_file(os.path.join(root, file))

    ui_count = len(ui_to_daos)
    logger.info(f"   Found {ui_count} UI components using DAOs")

    logger.info("\n[3/4] Building complete mapping...")

    # Build the complete mapping: StoredProc -> DAO -> UI
    complete_mapping = []

    for dao_name, stored_procs in dao_to_stored_procs.items():
        dao_file = dao_to_file.get(dao_name, 'Unknown')

        # Find all UI components that use this DAO
        ui_components_using_dao = []
        for ui_name, daos_used in ui_to_daos.items():
            if dao_name in daos_used:
                ui_components_using_dao.append({
                    'ui_name': ui_name,
                    'ui_type': ui_types.get(ui_name, 'Unknown'),
                    'controller_file': ui_to_file.get(ui_name, 'Unknown')
                })

        # Create mapping entries
        for sp_name in stored_procs:
            if ui_components_using_dao:
                # DAO is used by UI components
                for ui_info in ui_components_using_dao:
                    complete_mapping.append({
                        'Stored_Procedure': sp_name,
                        'DAO_Class': dao_name,
                        'DAO_File': dao_file,
                        'UI_Component': ui_info['ui_name'],
                        'UI_Type': ui_info['ui_type'],
                        'Controller_File': ui_info['controller_file']
                    })
            else:
                # DAO not used by any UI (or we couldn't find it)
                complete_mapping.append({
                    'Stored_Procedure': sp_name,
                    'DAO_Class': dao_name,
                    'DAO_File': dao_file,
                    'UI_Component': 'Not Found',
                    'UI_Type': 'N/A',
                    'Controller_File': 'N/A'
                })

    logger.info(f"   Created {len(complete_mapping)} complete mappings")

    logger.info("\n[4/4] Writing output files...")

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Write CSV - Complete mapping
    csv_file = os.path.join(output_dir, csv_filename)
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        fieldnames = ['Stored_Procedure', 'DAO_Class', 'DAO_File', 'UI_Component', 'UI_Type', 'Controller_File']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in sorted(complete_mapping, key=lambda x: (x['Stored_Procedure'], x['DAO_Class'], x['UI_Component'])):
            writer.writerow(row)

    logger.info(f"   CSV: {csv_file}")

    # Write Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "StoredProc to UI"

        headers = ['Stored Procedure', 'DAO Class', 'DAO File', 'UI Component', 'UI Type', 'UI File']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row in sorted(complete_mapping, key=lambda x: (x['Stored_Procedure'], x['DAO_Class'], x['UI_Component'])):
            ws.append([row['Stored_Procedure'], row['DAO_Class'], row['DAO_File'],
                       row['UI_Component'], row['UI_Type'], row['Controller_File']])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width

        excel_file = os.path.join(output_dir, excel_filename)
        wb.save(excel_file)
        logger.info(f"   Excel: {excel_file}")
    except ImportError:
        logger.info("   Excel: Skipped (openpyxl not available)")

    logger.info("\n" + "=" * 80)
    logger.info("SUMMARY")
    logger.info("=" * 80)
    logger.info(f"DAOs scanned: {dao_count}")
    logger.info(f"DAOs with stored procs: {len(dao_to_stored_procs)}")
    logger.info(f"UI components found: {ui_count}")
    logger.info(f"Total mappings: {len(complete_mapping)}")
    logger.info(f"Unique stored procedures: {len(set(m['Stored_Procedure'] for m in complete_mapping))}")
    logger.info("\nDone!")

if __name__ == "__main__":
    main()

