"""
Script to create a final Excel report with multiple tabs:
1. UI_Mappings_Final - from CSV file
2. Forward_Dependencies - from JSON file
3. Reverse_Dependencies - from JSON file

This consolidates all dependency analysis results into a single Excel workbook.
"""

import json
import csv
import os
import sys
import logging
from datetime import datetime
from config_loader import ConfigLoader

def load_csv_data(csv_file):
    """Load data from CSV file."""
    try:
        data = []
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for row in reader:
                data.append(row)
        return data, fieldnames
    except FileNotFoundError:
        raise FileNotFoundError(f"CSV file not found: {csv_file}")
    except csv.Error as e:
        raise ValueError(f"Error reading CSV file: {e}")

def load_json_data(json_file):
    """Load data from JSON file."""
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"JSON file not found: {json_file}")
    except json.JSONDecodeError as e:
        raise ValueError(f"Invalid JSON in file: {e}")

def flatten_json_results(json_data):
    """Flatten JSON data structure for Excel export."""
    flattened = []

    for item in json_data:
        procedure = item.get('procedure', 'Unknown')
        status = item.get('status', 'Unknown')

        if status == 'error':
            # Add error row
            flattened.append({
                'Database_Object': procedure,
                'Status': 'Error',
                'Error_Message': item.get('error', 'Unknown error'),
                'server_name': '',
                'object_name': '',
                'object_name_path': '',
                'referenced_object_fullname': '',
                'depth': '',
                'object_id_path': '',
                'object_type_desc_path': ''
            })
        else:
            results = item.get('results', [])
            if not results:
                # Add row indicating no dependencies found
                flattened.append({
                    'Database_Object': procedure,
                    'Status': 'No Dependencies',
                    'Error_Message': '',
                    'server_name': '',
                    'object_name': '',
                    'object_name_path': '',
                    'referenced_object_fullname': '',
                    'depth': '',
                    'object_id_path': '',
                    'object_type_desc_path': ''
                })
            else:
                # Add each result row
                for result in results:
                    row = {
                        'Database_Object': procedure,
                        'Status': 'Success',
                        'Error_Message': ''
                    }
                    # Add all fields from the result
                    row.update(result)
                    flattened.append(row)

    return flattened

def find_file_path(file_value, java_source_dirs):
    """
    Find the full path to a Java file in the source directories.

    Args:
        file_value: The file path from the CSV (e.g., "com/matrixcare/matrix/dao/CustomerDAO.java")
        java_source_dirs: List of Java source directories to search

    Returns:
        Full file path if found, None otherwise
    """
    if not file_value or file_value == '':
        return None

    # Try each source directory
    for source_dir in java_source_dirs:
        full_path = os.path.join(source_dir, file_value)
        if os.path.exists(full_path):
            return full_path

    return None

# Constants
MAX_COLUMN_WIDTH = 100

def auto_adjust_column_widths(worksheet, num_columns):
    """Auto-adjust column widths based on content."""
    from openpyxl.utils import get_column_letter

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def apply_borders_to_range(worksheet, min_row, max_row, min_col, max_col, border_style):
    """Apply border style to a range of cells."""
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border_style

def create_excel_report(csv_file, forward_json, reverse_json, output_file, java_source_dirs, log):
    """Create Excel workbook with multiple tabs."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl library not found. Please install it: pip install openpyxl")
        return False

    # Define border style for all cells
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    wb = None
    try:
        log.info("Creating Excel workbook...")
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
        # Tab 1: Controller Components
        log.info("  Adding tab 1: Controller_Components...")
        if os.path.exists(csv_file):
            csv_data, csv_headers = load_csv_data(csv_file)

            # Validate we have data
            if not csv_data:
                log.warning("    CSV file is empty - no data to add")

            # Validate required columns exist
            required_columns = ['DAO_File', 'Controller_File', 'Stored_Procedure', 'DAO_Class']
            missing_columns = [col for col in required_columns if col not in csv_headers]
            if missing_columns:
                log.warning(f"    Missing expected columns in CSV: {', '.join(missing_columns)}")

            ws1 = wb.create_sheet("Controller_Components")
        
            # Add headers
            ws1.append(list(csv_headers))

            # Style headers
            for cell in ws1[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Add data
            for row in csv_data:
                ws1.append([row.get(header, '') for header in csv_headers])

            # Find column indices for DAO_File and Controller_File
            dao_file_col = None
            controller_file_col = None
            for idx, header in enumerate(csv_headers, start=1):
                if header == 'DAO_File':
                    dao_file_col = idx
                elif header == 'Controller_File':
                    controller_file_col = idx

            # Add hyperlinks to DAO_File and Controller_File columns
            for row_idx in range(2, ws1.max_row + 1):
                # DAO_File hyperlink
                if dao_file_col:
                    cell = ws1.cell(row=row_idx, column=dao_file_col)
                    if cell.value:
                        file_path = find_file_path(cell.value, java_source_dirs)
                        if file_path:
                            cell.hyperlink = f"file:///{file_path}"
                            cell.font = Font(color="0563C1", underline="single")

                # Controller_File hyperlink
                if controller_file_col:
                    cell = ws1.cell(row=row_idx, column=controller_file_col)
                    if cell.value:
                        file_path = find_file_path(cell.value, java_source_dirs)
                        if file_path:
                            cell.hyperlink = f"file:///{file_path}"
                            cell.font = Font(color="0563C1", underline="single")

            # Apply borders to all cells (AFTER hyperlinks to preserve borders)
            apply_borders_to_range(ws1, 1, ws1.max_row, 1, len(csv_headers), thin_border)

            # Auto-adjust column widths
            auto_adjust_column_widths(ws1, len(csv_headers))

            # Add autofilter (dropdown boxes on headers)
            ws1.auto_filter.ref = ws1.dimensions

            log.info(f"    Added {len(csv_data)} rows with hyperlinks")
        else:
            log.warning(f"    CSV file not found: {csv_file}")
            ws1 = wb.create_sheet("Controller_Components")
            ws1.append(["File not found", csv_file])

        # Tab 2: DAO Components (distinct values from Controller_Components)
        log.info("  Adding tab 2: DAO_Components...")
        if os.path.exists(csv_file):
            # Extract distinct combinations of Stored_Procedure, DAO_Class, DAO_File
            dao_data_set = set()
            for row in csv_data:
                stored_proc = row.get('Stored_Procedure', '')
                dao_class = row.get('DAO_Class', '')
                dao_file = row.get('DAO_File', '')
                # Add as tuple to set (automatically handles uniqueness)
                dao_data_set.add((stored_proc, dao_class, dao_file))

            # Convert set to sorted list
            dao_data_list = sorted(list(dao_data_set))

            ws2 = wb.create_sheet("DAO_Components")

            # Add headers
            dao_headers = ['Stored_Procedure', 'DAO_Class', 'DAO_File']
            ws2.append(dao_headers)

            # Style headers
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Add data
            for dao_row in dao_data_list:
                ws2.append(list(dao_row))

            # Add hyperlinks to DAO_File column (column 3)
            dao_file_col = 3
            for row_idx in range(2, ws2.max_row + 1):
                cell = ws2.cell(row=row_idx, column=dao_file_col)
                if cell.value:
                    file_path = find_file_path(cell.value, java_source_dirs)
                    if file_path:
                        cell.hyperlink = f"file:///{file_path}"
                        cell.font = Font(color="0563C1", underline="single")

            # Apply borders to all cells (AFTER hyperlinks to preserve borders)
            apply_borders_to_range(ws2, 1, ws2.max_row, 1, len(dao_headers), thin_border)

            # Auto-adjust column widths
            auto_adjust_column_widths(ws2, len(dao_headers))

            # Add autofilter (dropdown boxes on headers)
            ws2.auto_filter.ref = ws2.dimensions

            log.info(f"    Added {len(dao_data_list)} distinct DAO component rows with hyperlinks")
        else:
            log.warning(f"    CSV file not found: {csv_file}")
            ws2 = wb.create_sheet("DAO_Components")
            ws2.append(["File not found", csv_file])

        # Tab 3: Forward Dependencies
        log.info("  Adding tab 3: Forward_Dependencies...")
        if os.path.exists(forward_json):
            forward_data = load_json_data(forward_json)

            # Validate we have data
            if not forward_data:
                log.warning("    Forward JSON file is empty - no data to add")

            forward_flattened = flatten_json_results(forward_data)
            ws3 = wb.create_sheet("Forward_Dependencies")

            # Define column order for Forward Dependencies
            if forward_flattened:
                headers = [
                    'Status',
                    'server_name',
                    'Database_Object',
                    'referenced_object_fullname',
                    'depth',
                    'object_name_path',
                    'object_id_path',
                    'object_type_desc_path',
                    'object_name',
                    'Error_Message'
                ]
                ws3.append(headers)

                # Style headers
                for cell in ws3[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

                # Add data
                for row in forward_flattened:
                    ws3.append([row.get(header, '') for header in headers])

                # Apply borders to all cells
                apply_borders_to_range(ws3, 1, ws3.max_row, 1, len(headers), thin_border)

                # Auto-adjust column widths
                auto_adjust_column_widths(ws3, len(headers))

                # Add autofilter (dropdown boxes on headers)
                ws3.auto_filter.ref = ws3.dimensions

                log.info(f"    Added {len(forward_flattened)} rows")
            else:
                ws3.append(["No data available"])
                log.info("    No data to add")
        else:
            log.warning(f"    JSON file not found: {forward_json}")
            ws3 = wb.create_sheet("Forward_Dependencies")
            ws3.append(["File not found", forward_json])

        # Tab 4: Reverse Dependencies
        log.info("  Adding tab 4: Reverse_Dependencies...")
        if os.path.exists(reverse_json):
            reverse_data = load_json_data(reverse_json)

            # Validate we have data
            if not reverse_data:
                log.warning("    Reverse JSON file is empty - no data to add")

            reverse_flattened = flatten_json_results(reverse_data)
            ws4 = wb.create_sheet("Reverse_Dependencies")

            # Define column order for Reverse Dependencies
            if reverse_flattened:
                headers = [
                    'Status',
                    'server_name',
                    'Database_Object',
                    'referencing_object_fullname',
                    'depth',
                    'object_name_path',
                    'object_id_path',
                    'object_type_desc_path',
                    'object_name',
                    'Error_Message'
                ]
                ws4.append(headers)

                # Style headers
                for cell in ws4[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

                # Add data
                for row in reverse_flattened:
                    ws4.append([row.get(header, '') for header in headers])

                # Apply borders to all cells
                apply_borders_to_range(ws4, 1, ws4.max_row, 1, len(headers), thin_border)

                # Auto-adjust column widths
                auto_adjust_column_widths(ws4, len(headers))

                # Add autofilter (dropdown boxes on headers)
                ws4.auto_filter.ref = ws4.dimensions

                log.info(f"    Added {len(reverse_flattened)} rows")
            else:
                ws4.append(["No data available"])
                log.info("    No data to add")
        else:
            log.warning(f"    JSON file not found: {reverse_json}")
            ws4 = wb.create_sheet("Reverse_Dependencies")
            ws4.append(["File not found", reverse_json])

        # Save workbook
        log.info(f"Saving Excel file to: {output_file}")
        wb.save(output_file)
        log.info("Excel file created successfully!")

        return True

    except Exception as e:
        log.error(f"Error creating Excel report: {str(e)}")
        import traceback
        log.error(traceback.format_exc())
        return False

def main():
    # Load configuration
    config = ConfigLoader()

    # Get configuration values
    base_dir = config.get_project_base_dir()
    output_dir = config.get_output_dir()
    log_dir = config.get_log_dir()

    # Get Java source directories for hyperlinks
    java_source_dirs = config.get_java_source_dirs()

    # Define file paths
    csv_file = os.path.join(output_dir, config.get_final_ui_mappings_csv())
    forward_json = os.path.join(output_dir, config.get_object_dependency_list_forward_json())
    reverse_json = os.path.join(output_dir, config.get_object_dependency_list_reverse_json())
    output_file = os.path.join(output_dir, config.get_final_excel_report())

    # Setup logging
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    log_filename = f'log_06_create_final_excel_file_{timestamp}.log'
    log_filepath = os.path.join(log_dir, log_filename)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filepath, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    logging.info("="*60)
    logging.info("Creating Final Excel Report")
    logging.info("="*60)
    logging.info(f"CSV Input:     {csv_file}")
    logging.info(f"Forward JSON:  {forward_json}")
    logging.info(f"Reverse JSON:  {reverse_json}")
    logging.info(f"Excel Output:  {output_file}")
    logging.info(f"Java Source Directories:")
    for src_dir in java_source_dirs:
        logging.info(f"  - {src_dir}")
    logging.info("")

    # Create Excel report
    success = create_excel_report(csv_file, forward_json, reverse_json, output_file, java_source_dirs, logging)

    if success:
        logging.info("")
        logging.info("="*60)
        logging.info("COMPLETE!")
        logging.info("="*60)
        logging.info(f"Excel file: {output_file}")
        logging.info(f"Log file:   {log_filepath}")
    else:
        logging.error("Failed to create Excel report")
        sys.exit(1)

if __name__ == "__main__":
    main()


