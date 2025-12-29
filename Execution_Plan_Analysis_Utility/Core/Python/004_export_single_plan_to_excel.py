import json
import pandas as pd
from pathlib import Path
import sys
import logging
from datetime import datetime
import configparser
import subprocess

# ============================================================================
# CONFIGURATION AND LOGGING SETUP
# ============================================================================

# Calculate base directory dynamically (project root)
# Current structure: ProjectRoot\Core\Python\004_export_single_plan_to_excel.py
# So we need to go up 2 levels: Python -> Core -> ProjectRoot
base_dir = Path(__file__).parent.parent.parent

# Load the INI configuration file
ini_config = configparser.ConfigParser(interpolation=configparser.ExtendedInterpolation())
config_file = Path(__file__).parent / 'config.ini'
ini_config.read(config_file, encoding='utf-8')

# Create logs directory if it doesn't exist (relative to base_dir)
logs_dir = base_dir / ini_config['Paths']['logs_dir']
logs_dir.mkdir(parents=True, exist_ok=True)

# Create log filename with timestamp
timestamp_format = ini_config['Logging']['timestamp_format']
timestamp = datetime.now().strftime(timestamp_format)
log_filename = ini_config['Logging']['single_plan_export_log_file']
LOG_FILE = logs_dir / f'{log_filename}_{timestamp}.log'

# Configure logging
log_format = ini_config['Logging']['log_format']
log_level = getattr(logging, ini_config['Logging']['log_level'])
log_filemode = ini_config.get('Logging', 'log_filemode', fallback='w')

logging.basicConfig(
    filename=str(LOG_FILE),
    level=log_level,
    format=log_format,
    filemode=log_filemode
)


def load_json_data(json_file: str) -> dict:
    """Load the execution plan analysis JSON file."""
    # Use utf-8-sig to handle BOM if present
    with open(json_file, 'r', encoding='utf-8-sig') as f:
        return json.load(f)


def create_summary_sheet(plans: list) -> pd.DataFrame:
    """Create summary sheet with metrics for all plans (no comparison)."""

    rows = []

    # Define metrics to display
    metrics = [
        ('total_estimated_cost', 'Total Estimated Cost'),
        ('total_elapsed_time_ms', 'Total Elapsed Time (ms)'),
        ('total_cpu_time_ms', 'Total CPU Time (ms)'),
        ('total_wait_time_ms', 'Total Wait Time (ms)'),
        ('total_logical_reads', 'Total Logical Reads'),
        ('optimizer_timeouts', 'Optimizer Timeouts'),
        ('total_statements', 'Total Statements')
    ]

    for metric_key, metric_label in metrics:
        row = {'Metric': metric_label}

        for plan_data in plans:
            plan_name = plan_data.get('config_name', plan_data['plan_name'])
            summary = plan_data['summary']

            # Get the metric value
            if metric_key == 'total_wait_time_ms':
                # Calculate wait time if not present
                value = summary.get(metric_key, 0)
            else:
                value = summary.get(metric_key, 0)

            row[plan_name] = value

        rows.append(row)

    return pd.DataFrame(rows)


def create_plan_overview_sheet(plans: list) -> pd.DataFrame:
    """Create plan overview sheet with high-level summary."""

    rows = []

    for plan_data in plans:
        summary = plan_data['summary']

        row = {
            'Plan Name': plan_data.get('config_name', plan_data['plan_name']),
            'Total Statements': summary['total_statements'],
            'Total Estimated Cost': summary['total_estimated_cost'],
            'Total Elapsed Time (ms)': summary['total_elapsed_time_ms'],
            'Total CPU Time (ms)': summary['total_cpu_time_ms'],
            'Total Wait Time (ms)': summary.get('total_wait_time_ms', 0),
            'Total Logical Reads': summary['total_logical_reads'],
            'Optimizer Timeouts': summary['optimizer_timeouts'],
            'Missing Indexes Count': len(summary.get('missing_indexes', []))
        }
        rows.append(row)

    return pd.DataFrame(rows)


def create_statements_sheet(plan_data: dict, plan_name: str) -> pd.DataFrame:
    """Create detailed statements sheet for a plan."""

    rows = []
    for stmt in plan_data['statements']:
        row = {
            'Statement ID': stmt['statement_id'],
            'Type': stmt['statement_type'],
            'Early Abort Reason': stmt.get('early_abort_reason') if stmt.get('early_abort_reason') else '',
            'Query Preview': stmt['statement_text_preview'],
            'Missing Indexes': len(stmt.get('missing_indexes', [])),
            'Estimated Cost': stmt['estimated_cost'],
            'Estimated Rows': stmt['estimated_rows'],
            'Actual Rows': stmt.get('actual_rows', 0),
            'Elapsed Time (ms)': stmt['elapsed_time_ms'],
            'CPU Time (ms)': stmt['cpu_time_ms'],
            'Wait Time (ms)': stmt.get('wait_time_ms', 0),
            'Logical Reads': stmt['logical_reads'],
            'Actual Executions': stmt.get('actual_executions', 0),
            'Optimizer Level': stmt['optimizer_level']
        }
        rows.append(row)

    return pd.DataFrame(rows)


def create_missing_indexes_sheet(plans: list) -> pd.DataFrame:
    """Create missing indexes sheet."""

    rows = []

    for plan_data in plans:
        plan_name = plan_data.get('config_name', plan_data['plan_name'])

        for stmt in plan_data['statements']:
            for idx in stmt.get('missing_indexes', []):
                row = {
                    'Plan': plan_name,
                    'Statement ID': stmt['statement_id'],
                    'Impact %': idx['impact_percent'],
                    'Database': idx['database'],
                    'Schema': idx['schema'],
                    'Table': idx['table'],
                    'Equality Columns': ', '.join(idx['equality_columns']),
                    'Inequality Columns': ', '.join(idx['inequality_columns']),
                    'Include Columns': ', '.join(idx['include_columns'])
                }
                rows.append(row)

    if not rows:
        # Return empty dataframe with columns if no missing indexes
        return pd.DataFrame(columns=[
            'Plan', 'Statement ID', 'Impact %', 'Database', 'Schema', 'Table',
            'Equality Columns', 'Inequality Columns', 'Include Columns'
        ])

    df = pd.DataFrame(rows)
    df = df.sort_values('Impact %', ascending=False)
    return df


def create_warnings_sheet(plans: list) -> pd.DataFrame:
    """Create warnings sheet combining warnings from all plans."""

    rows = []

    for plan_data in plans:
        plan_name = plan_data.get('config_name', plan_data['plan_name'])

        # Get warnings from the plan
        warnings = plan_data.get('warnings', [])

        for warning in warnings:
            warning_type = warning.get('type', 'Unknown')
            statement_id = warning.get('statement_id', '')

            # Build description from all attributes (excluding type and statement_id)
            attributes = {k: v for k, v in warning.items() if k not in ['type', 'statement_id']}
            description = ', '.join([f"{k}: {v}" for k, v in attributes.items()]) if attributes else 'No additional details'

            row = {
                'Plan': plan_name,
                'Statement ID': statement_id,
                'Warning Type': warning_type,
                'Description': description
            }
            rows.append(row)

    if not rows:
        # Return empty dataframe with columns if no warnings
        return pd.DataFrame(columns=['Plan', 'Statement ID', 'Warning Type', 'Description'])

    return pd.DataFrame(rows)


def create_node_details_sheet(plans: list) -> pd.DataFrame:
    """Create node details sheet with detailed information for each operator node."""

    def remove_brackets(text):
        """Remove square brackets from text."""
        if not text:
            return ''
        # First remove empty brackets followed by period: [].
        result = str(text).replace('[].', '')
        # Then remove remaining [ and ]
        result = result.replace('[', '').replace(']', '')
        return result

    def extract_table_index_name(full_path):
        """Extract just the table/index name from a three-part naming convention.

        Examples:
            'QA07_Greg.dbo.AR_ChargeRecords.CIDX_AR_ChargeRecords_PatientID' -> 'CIDX_AR_ChargeRecords_PatientID'
            'dbo.TableName' -> 'TableName'
            'SimpleTable' -> 'SimpleTable'
        """
        if not full_path:
            return ''

        # Split by period to check for three-part naming
        parts = full_path.split('.')

        # If we have 3 or more parts (database.schema.table.index or database.schema.table)
        # Return the last part (table or index name)
        if len(parts) >= 3:
            return parts[-1]

        # Otherwise, return the original value
        return full_path

    rows = []

    for plan in plans:
        plan_name = plan.get('plan_name', 'Unknown')

        for stmt in plan.get('statements', []):
            # Get node details if available
            node_details = stmt.get('node_details', [])

            for node in node_details:
                full_table_index = remove_brackets(node.get('table_index', ''))
                short_table_index = extract_table_index_name(full_table_index)

                row = {
                    'Statement ID': node.get('statement_id', ''),
                    'Node ID': node.get('node_id', ''),
                    'Node Type': node.get('node_type', ''),
                    'Table/Index Name': short_table_index,
                    'Seek Predicates': remove_brackets(node.get('seek_predicates', '')),
                    'Predicate': remove_brackets(node.get('predicate', '')),
                    'Output List': remove_brackets(node.get('output_list', '')),
                    'Physical Op': node.get('physical_op', ''),
                    'Logical Op': node.get('logical_op', ''),
                    'Est. Cost': node.get('estimated_cost', 0),
                    'Est. CPU Cost': node.get('estimated_cpu_cost', 0),
                    'Est. I/O Cost': node.get('estimated_io_cost', 0),
                    'Est. Executions': node.get('estimated_executions', 0),
                    'Est. Rows': node.get('estimated_rows', 0),
                    'Actual Rows': node.get('actual_rows', 0),
                    'Actual Executions': node.get('actual_executions', 0),
                    'Actual Rebinds': node.get('actual_rebinds', 0),
                    'Actual Rewinds': node.get('actual_rewinds', 0),
                    'Parallel': node.get('parallel', 'No'),
                    'Warnings': node.get('warnings', ''),
                    'Table/Index Full Path': full_table_index
                }
                rows.append(row)

    if not rows:
        # Return empty DataFrame with correct columns
        return pd.DataFrame(columns=[
            'Statement ID', 'Node ID', 'Node Type', 'Table/Index Name',
            'Seek Predicates', 'Predicate', 'Output List', 'Physical Op', 'Logical Op',
            'Est. Cost', 'Est. CPU Cost', 'Est. I/O Cost', 'Est. Executions',
            'Est. Rows', 'Actual Rows', 'Actual Executions', 'Actual Rebinds', 'Actual Rewinds',
            'Parallel', 'Warnings', 'Table/Index Full Path'
        ])

    return pd.DataFrame(rows)


def format_excel_file(excel_file: str, ini_config: configparser.ConfigParser):
    """Apply formatting to the Excel file using openpyxl."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(excel_file)

        # Define styles from config
        header_color = ini_config['Excel']['header_color']
        header_font_color = ini_config['Excel']['header_font_color']
        header_font_size = int(ini_config['Excel']['header_font_size'])

        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        header_font = Font(bold=True, color=header_font_color, size=header_font_size)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Light red fill for 10x cardinality estimation errors
        light_red_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Apply borders to all cells (including blank cells)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border

            # Apply 10x rule highlighting for Statements sheets
            if 'Statements' in sheet_name:
                # Find column indices for Estimated Rows and Actual Rows
                header_row = [cell.value for cell in ws[1]]
                try:
                    est_rows_col = header_row.index('Estimated Rows') + 1
                    act_rows_col = header_row.index('Actual Rows') + 1

                    # Iterate through data rows (skip header)
                    for row_idx in range(2, ws.max_row + 1):
                        est_rows_cell = ws.cell(row=row_idx, column=est_rows_col)
                        act_rows_cell = ws.cell(row=row_idx, column=act_rows_col)

                        est_rows = est_rows_cell.value
                        act_rows = act_rows_cell.value

                        # Convert to numbers, handle None/empty values
                        try:
                            est_rows = float(est_rows) if est_rows is not None else 0
                            act_rows = float(act_rows) if act_rows is not None else 0
                        except (ValueError, TypeError):
                            continue

                        # Apply 10x rule highlighting
                        highlight = False

                        # Check for 10x difference
                        has_10x_difference = False
                        if est_rows > 0 and act_rows >= 10 * est_rows:
                            has_10x_difference = True  # Underestimation
                        elif act_rows > 0 and est_rows >= 10 * act_rows:
                            has_10x_difference = True  # Overestimation
                        elif est_rows > 0 and act_rows == 0:
                            has_10x_difference = True  # Overestimation (got nothing)
                        elif act_rows > 0 and est_rows == 0:
                            has_10x_difference = True  # Underestimation (expected nothing)

                        # Only highlight if there's a 10x difference AND at least one value > 200
                        if has_10x_difference and (est_rows > 200 or act_rows > 200):
                            highlight = True

                        if highlight:
                            est_rows_cell.fill = light_red_fill
                            act_rows_cell.fill = light_red_fill

                except ValueError:
                    # Columns not found in this sheet, skip highlighting
                    pass

            # Apply 10x rule highlighting for Node Details sheets
            if 'Node Details' in sheet_name:
                # Find column indices for Est. Rows and Actual Rows
                header_row = [cell.value for cell in ws[1]]
                try:
                    est_rows_col = header_row.index('Est. Rows') + 1
                    act_rows_col = header_row.index('Actual Rows') + 1

                    # Iterate through data rows (skip header)
                    for row_idx in range(2, ws.max_row + 1):
                        est_rows_cell = ws.cell(row=row_idx, column=est_rows_col)
                        act_rows_cell = ws.cell(row=row_idx, column=act_rows_col)

                        est_rows = est_rows_cell.value
                        act_rows = act_rows_cell.value

                        # Convert to numbers, handle None/empty values
                        try:
                            est_rows = float(est_rows) if est_rows is not None else 0
                            act_rows = float(act_rows) if act_rows is not None else 0
                        except (ValueError, TypeError):
                            continue

                        # Apply 10x rule highlighting
                        highlight = False

                        # Check for 10x difference
                        has_10x_difference = False
                        if est_rows > 0 and act_rows >= 10 * est_rows:
                            has_10x_difference = True  # Underestimation
                        elif act_rows > 0 and est_rows >= 10 * act_rows:
                            has_10x_difference = True  # Overestimation
                        elif est_rows > 0 and act_rows == 0:
                            has_10x_difference = True  # Overestimation (got nothing)
                        elif act_rows > 0 and est_rows == 0:
                            has_10x_difference = True  # Underestimation (expected nothing)

                        # Only highlight if there's a 10x difference AND at least one value > 200
                        if has_10x_difference and (est_rows > 200 or act_rows > 200):
                            highlight = True

                        if highlight:
                            est_rows_cell.fill = light_red_fill
                            act_rows_cell.fill = light_red_fill

                except ValueError:
                    # Columns not found in this sheet, skip highlighting
                    pass

            # Add autofilter to header row
            if ws.max_row > 0 and ws.max_column > 0:
                ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

            # Freeze panes (header row)
            ws.freeze_panes = 'A2'

            # Auto-adjust column widths (after filter is applied)
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                    except:
                        pass

                # Add extra space for filter dropdown arrow in header (3 extra chars minimum)
                # Use at least 10 for very short columns to accommodate filter button
                adjusted_width = max(10, min(max_length + 3, 100))
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(excel_file)
        logging.info(f"✅ Excel formatting applied successfully!")

    except ImportError:
        logging.warning("openpyxl not installed. Excel file created without formatting.")
        logging.info("   Install with: pip install openpyxl")
    except Exception as e:
        logging.warning(f"Could not apply formatting: {e}")


def main():
    """Main function to export individual plans to Excel."""

    logging.info("="*80)
    logging.info("EXPORT INDIVIDUAL PLANS TO EXCEL - STARTED")
    logging.info("="*80)
    logging.info(f"Log file: {LOG_FILE}")

    # Load analysis results
    output_dir = base_dir / ini_config['Paths']['output_dir']
    json_file = output_dir / 'execution_plan_single_analysis_json.json'

    if not json_file.exists():
        logging.error(f"Analysis file not found: {json_file}")
        logging.error("Please run 003_analyze_single_plan.py first!")
        sys.exit(1)

    logging.info(f"Loading analysis data from {json_file}...")
    data = load_json_data(str(json_file))

    analysis_timestamp = data['analysis_timestamp']
    total_plans = data['total_plans']
    plans = data['plans']

    logging.info(f"Analysis timestamp: {analysis_timestamp}")
    logging.info(f"Total plans to export: {total_plans}")

    # Create timestamp for filenames
    file_timestamp = datetime.now().strftime(ini_config['Logging']['timestamp_format'])

    # Get config settings
    excel_engine = ini_config['Excel']['excel_engine']
    max_sheet_name_length = int(ini_config['Excel']['max_sheet_name_length'])

    # Create one Excel file per plan
    excel_files = []

    for idx, plan_data in enumerate(plans, 1):
        # Get the plan name and original filename
        plan_name = plan_data.get('config_name', plan_data['plan_name'])
        safe_plan_name = "".join(c for c in plan_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')

        file_path = plan_data.get('file_path', '')
        if file_path:
            original_filename = Path(file_path).stem  # Gets filename without extension
        else:
            # Fallback to plan name if file path not available
            original_filename = safe_plan_name

        # Create output filename: Summary.<plan name>.<file name>.<date>.<time>.xlsx
        # Split timestamp into date and time parts (format: YYYYMMDD_HHMMSS)
        date_part = file_timestamp[:8]  # YYYYMMDD
        time_part = file_timestamp[9:]  # HHMMSS
        excel_file = output_dir / f'Summary.{safe_plan_name}.{original_filename}.{date_part}.{time_part}.xlsx'

        logging.info(f"\n[{idx}/{total_plans}] Creating Excel file for: {plan_data['plan_name']}")
        logging.info(f"  Output file: {excel_file.name}")

        # Wrap plan in a list for the sheet creation functions
        plan_list = [plan_data]

        with pd.ExcelWriter(str(excel_file), engine=excel_engine) as writer:

            # Sheet 1: Summary (without Winner/Difference columns)
            logging.info("    Creating Summary sheet...")
            df_summary = create_summary_sheet(plan_list)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)

            # Sheet 2: Plan Overview
            logging.info("    Creating Plan Overview sheet...")
            df_overview = create_plan_overview_sheet(plan_list)
            df_overview.to_excel(writer, sheet_name='Plan Overview', index=False)

            # Sheet 3: Plan Statements
            plan_name = plan_data.get('config_name', plan_data['plan_name'])
            logging.info(f"    Creating {plan_name} Statements sheet...")
            df_plan_stmts = create_statements_sheet(plan_data, plan_name)
            df_plan_stmts.to_excel(writer, sheet_name='Statements', index=False)

            # Sheet 4: Node Details
            logging.info("    Creating Details sheet...")
            df_node_details = create_node_details_sheet(plan_list)
            df_node_details.to_excel(writer, sheet_name='Details', index=False)

            # Sheet 5: Missing Indexes
            logging.info("    Creating Missing Indexes sheet...")
            df_missing = create_missing_indexes_sheet(plan_list)
            df_missing.to_excel(writer, sheet_name='Missing Indexes', index=False)

            # Sheet 6: Warnings
            logging.info("    Creating Warnings sheet...")
            df_warnings = create_warnings_sheet(plan_list)
            df_warnings.to_excel(writer, sheet_name='Warnings', index=False)

        logging.info(f"  ✅ Excel file created: {excel_file.name}")

        # Apply formatting
        logging.info("    Applying formatting...")
        format_excel_file(str(excel_file), ini_config)

        excel_files.append(excel_file)

    logging.info("\n" + "="*80)
    logging.info("EXPORT COMPLETE")
    logging.info("="*80)
    logging.info(f"\nTotal Excel files created: {len(excel_files)}")

    for excel_file in excel_files:
        logging.info(f"  - {excel_file.name}")

    # Create a completion flag file for PowerShell to detect
    completion_flag = output_dir / '.single_plan_complete'
    try:
        with open(completion_flag, 'w', encoding='utf-8') as f:
            f.write('COMPLETE')
        logging.info(f"\nCompletion flag created: {completion_flag}")
    except Exception as e:
        logging.warning(f"Could not create completion flag: {e}")

    logging.info("\n" + "="*80)


if __name__ == '__main__':
    main()


