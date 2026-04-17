import json
import pandas as pd
from pathlib import Path
import os
import sys
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Any, Set
from config_loader import ConfigLoader
from excel_export_helpers import (
    sanitize_filename, create_unique_sheet_name, parse_timestamp,
    apply_cardinality_highlighting, format_worksheet, open_excel_file, set_tab_colors
)

# ============================================================================
# CONFIGURATION AND LOGGING SETUP
# ============================================================================

# Load configuration
try:
    config = ConfigLoader()
except FileNotFoundError as e:
    print(f"[ERROR] {e}")
    print("Please ensure config.json exists in the Config directory.")
    sys.exit(1)
except Exception as e:
    print(f"[ERROR] Configuration error: {e}")
    print("Please check your config.json file for errors.")
    sys.exit(1)

# Get base directory
base_dir = config.get_base_dir()

# Setup logging
LOG_FILE = config.setup_logging(config.get_export_log_file())


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def load_json_data(json_file: str) -> dict:
    """
    Load the execution plan analysis JSON file.

    Args:
        json_file: Path to JSON file

    Returns:
        Dictionary containing execution plan data

    Raises:
        FileNotFoundError: If JSON file doesn't exist
        json.JSONDecodeError: If JSON is malformed
    """
    if not Path(json_file).exists():
        raise FileNotFoundError(f"JSON file not found: {json_file}")

    try:
        # Use utf-8-sig to handle BOM if present
        with open(json_file, 'r', encoding='utf-8-sig') as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        raise json.JSONDecodeError(f"Invalid JSON in {json_file}: {e}", e.doc, e.pos)


def validate_json_structure(data: dict) -> Tuple[bool, str]:
    """
    Validate that JSON has required structure.

    Args:
        data: JSON data dictionary

    Returns:
        Tuple of (is_valid, error_message)
    """
    if not isinstance(data, dict):
        return False, "Data is not a dictionary"

    if 'comparison' not in data:
        return False, "Missing 'comparison' key in data"

    # Check for either plan1/plan2 or version_1/version_greg structure
    has_plan_structure = 'plan1' in data and 'plan2' in data
    has_version_structure = 'version_1' in data and 'version_greg' in data

    if not (has_plan_structure or has_version_structure):
        return False, "Missing plan data (expected 'plan1'/'plan2' or 'version_1'/'version_greg')"

    return True, ""


def get_plan_keys(data: dict) -> Tuple[str, str]:
    """
    Determine which plan keys to use based on data structure.

    Args:
        data: JSON data dictionary

    Returns:
        Tuple of (plan1_key, plan2_key)
    """
    if 'plan1' in data and 'plan2' in data:
        return 'plan1', 'plan2'
    elif 'version_1' in data:
        # Use more generic approach - find all keys that aren't 'comparison' or metadata
        plan_keys = [k for k in data.keys() if k not in ['comparison', 'analysis_timestamp', 'config_file']]
        if len(plan_keys) >= 2:
            return plan_keys[0], plan_keys[1]
        return 'version_1', 'version_greg'
    else:
        # Fallback - find first two plan-like keys
        plan_keys = [k for k in data.keys() if k not in ['comparison', 'analysis_timestamp', 'config_file']]
        if len(plan_keys) >= 2:
            return plan_keys[0], plan_keys[1]
        raise ValueError("Could not determine plan keys from data structure")


def safe_get_plan_data(data: dict, plan_key: str) -> Optional[dict]:
    """
    Safely get plan data with validation.

    Args:
        data: JSON data dictionary
        plan_key: Key for the plan to retrieve

    Returns:
        Plan data dictionary or None if not found
    """
    if plan_key not in data:
        logging.warning(f"Plan key '{plan_key}' not found in data")
        return None
    return data[plan_key]


def create_summary_sheet(data: dict) -> pd.DataFrame:
    """
    Create summary comparison sheet.

    Args:
        data: JSON data dictionary

    Returns:
        DataFrame with metric comparison
    """
    comparison = data.get('comparison', {})
    if not comparison:
        logging.warning("No comparison data found")
        return pd.DataFrame()

    rows = []
    for metric, values in comparison.get('metrics', {}).items():
        plan1_name = comparison.get('plan1_name', 'Plan 1')
        plan2_name = comparison.get('plan2_name', 'Plan 2')

        # Handle both numeric and 'N/A' percent differences
        pct_diff = values.get('percent_difference', 0)
        if isinstance(pct_diff, str):
            pct_diff_display = pct_diff
        else:
            pct_diff_display = pct_diff

        row = {
            'Metric': metric.replace('_', ' ').title(),
            plan1_name: values.get(plan1_name, 0),
            plan2_name: values.get(plan2_name, 0),
            'Winner': values.get('winner', 'tie'),
            'Difference %': pct_diff_display,
            'Weight': values.get('weight', 'N/A')
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    return df


def create_statements_sheet(plan_data: dict, plan_name: str) -> pd.DataFrame:
    """Create detailed statements sheet for a plan."""

    rows = []
    for stmt in plan_data['statements']:
        row = {
            'Statement ID': stmt['statement_id'],
            'Type': stmt['statement_type'],
            'Early Abort Reason': stmt.get('early_abort_reason') if stmt.get('early_abort_reason') else '',
            'Query Preview': stmt['statement_text_preview'],
            'Missing Indexes': len(stmt.get('missing_indexes', [])),
            'Estimated Cost': stmt['estimated_cost'],
            'Estimated Rows': stmt['estimated_rows'],
            'Actual Rows': stmt.get('actual_rows', 0),
            'Elapsed Time (ms)': stmt['elapsed_time_ms'],
            'CPU Time (ms)': stmt['cpu_time_ms'],
            'Wait Time (ms)': stmt.get('wait_time_ms', 0),
            'Logical Reads': stmt['logical_reads'],
            'Actual Executions': stmt.get('actual_executions', 0),
            'Optimizer Level': stmt['optimizer_level']
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    return df


def create_missing_indexes_sheet(data: dict) -> pd.DataFrame:
    """Create missing indexes sheet."""

    rows = []

    # Handle both old and new JSON structure
    plan_keys = ['plan1', 'plan2'] if 'plan1' in data else ['version_1', 'version_greg']

    for plan_key in plan_keys:
        if plan_key not in data:
            continue

        plan_data = data[plan_key]

        for stmt in plan_data['statements']:
            for idx in stmt['missing_indexes']:
                row = {
                    'Plan': plan_data.get('config_name', plan_data['plan_name']),
                    'Statement ID': stmt['statement_id'],
                    'Impact %': idx['impact_percent'],
                    'Database': idx['database'],
                    'Schema': idx['schema'],
                    'Table': idx['table'],
                    'Equality Columns': ', '.join(idx['equality_columns']),
                    'Inequality Columns': ', '.join(idx['inequality_columns']),
                    'Include Columns': ', '.join(idx['include_columns'])
                }
                rows.append(row)

    if not rows:
        # Return empty dataframe with columns if no missing indexes
        return pd.DataFrame(columns=[
            'Plan', 'Statement ID', 'Impact %', 'Database', 'Schema', 'Table',
            'Equality Columns', 'Inequality Columns', 'Include Columns'
        ])

    df = pd.DataFrame(rows)
    df = df.sort_values('Impact %', ascending=False)
    return df


def create_plan_overview_sheet(data: dict) -> pd.DataFrame:
    """Create plan overview sheet with high-level summary."""

    rows = []

    # Handle both old and new JSON structure
    plan_keys = ['plan1', 'plan2'] if 'plan1' in data else ['version_1', 'version_greg']

    for plan_key in plan_keys:
        if plan_key not in data:
            continue

        plan_data = data[plan_key]
        summary = plan_data['summary']

        row = {
            'Plan Name': plan_data.get('config_name', plan_data['plan_name']),
            'Total Statements': summary['total_statements'],
            'Total Estimated Cost': summary['total_estimated_cost'],
            'Total Elapsed Time (ms)': summary['total_elapsed_time_ms'],
            'Total CPU Time (ms)': summary['total_cpu_time_ms'],
            'Total Wait Time (ms)': summary['total_wait_time_ms'],
            'Total Logical Reads': summary['total_logical_reads'],
            'Optimizer Timeouts': summary['optimizer_timeouts'],
            'Missing Indexes Count': len(summary['missing_indexes'])
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    return df


def create_winner_analysis_sheet(data: dict) -> pd.DataFrame:
    """
    Create winner analysis sheet with weighted scoring.

    Args:
        data: JSON data dictionary

    Returns:
        DataFrame with winner analysis
    """
    comparison = data.get('comparison', {})
    if not comparison:
        return pd.DataFrame()

    rows = [
        {'Analysis': 'Overall Winner', 'Value': comparison.get('winner', 'N/A')},
        {'Analysis': '', 'Value': ''}
    ]

    # Weighted scores removed - no longer displayed in Summary sheet

    rows.append({'Analysis': 'Winner Reasons:', 'Value': ''})

    for reason in comparison.get('winner_reasons', []):
        rows.append({'Analysis': '  • ' + reason, 'Value': ''})

    df = pd.DataFrame(rows)
    return df


def create_warnings_sheet(data: dict) -> pd.DataFrame:
    """Create warnings sheet combining warnings from both plans."""

    rows = []

    # Handle both old and new JSON structure
    plan_keys = ['plan1', 'plan2'] if 'plan1' in data else ['version_1', 'version_greg']

    for plan_key in plan_keys:
        if plan_key not in data:
            continue

        plan_data = data[plan_key]
        plan_name = plan_data.get('config_name', plan_data['plan_name'])

        # Get warnings from the plan
        warnings = plan_data.get('warnings', [])

        for warning in warnings:
            warning_type = warning.get('type', 'Unknown')
            statement_id = warning.get('statement_id', '')

            # Build description from all attributes (excluding type and statement_id)
            attributes = {k: v for k, v in warning.items() if k not in ['type', 'statement_id']}
            description = ', '.join([f"{k}: {v}" for k, v in attributes.items()]) if attributes else 'No additional details'

            row = {
                'Plan': plan_name,
                'Statement ID': statement_id,
                'Warning Type': warning_type,
                'Description': description
            }
            rows.append(row)

    if not rows:
        # Return empty dataframe with columns if no warnings
        return pd.DataFrame(columns=['Plan', 'Statement ID', 'Warning Type', 'Description'])

    df = pd.DataFrame(rows)
    return df


# Weighted Score sheet removed - function deleted


def create_parameters_sheet(data: dict) -> pd.DataFrame:
    """
    Create parameters comparison sheet showing compiled and runtime parameter values.

    Args:
        data: JSON data dictionary

    Returns:
        DataFrame with parameter information from both plans
    """
    plan1_data = data.get('plan1', {})
    plan2_data = data.get('plan2', {})
    plan1_name = data.get('comparison', {}).get('plan1_name', 'Plan 1')
    plan2_name = data.get('comparison', {}).get('plan2_name', 'Plan 2')

    rows = []

    # Collect all parameters from both plans
    plan1_params = {}
    plan2_params = {}

    # Extract parameters from Plan 1
    for stmt in plan1_data.get('statements', []):
        for param in stmt.get('parameters', []):
            param_name = param.get('parameter_name', '')
            if param_name and param_name not in plan1_params:
                plan1_params[param_name] = param

    # Extract parameters from Plan 2
    for stmt in plan2_data.get('statements', []):
        for param in stmt.get('parameters', []):
            param_name = param.get('parameter_name', '')
            if param_name and param_name not in plan2_params:
                plan2_params[param_name] = param

    # Get all unique parameter names
    all_param_names = sorted(set(list(plan1_params.keys()) + list(plan2_params.keys())))

    # Build comparison rows
    for param_name in all_param_names:
        plan1_param = plan1_params.get(param_name, {})
        plan2_param = plan2_params.get(param_name, {})

        rows.append({
            'Parameter Name': param_name,
            f'{plan1_name} - Compiled Value': plan1_param.get('compiled_value', 'N/A'),
            f'{plan1_name} - Runtime Value': plan1_param.get('runtime_value', 'N/A'),
            f'{plan1_name} - Data Type': plan1_param.get('data_type', 'N/A'),
            f'{plan2_name} - Compiled Value': plan2_param.get('compiled_value', 'N/A'),
            f'{plan2_name} - Runtime Value': plan2_param.get('runtime_value', 'N/A'),
            f'{plan2_name} - Data Type': plan2_param.get('data_type', 'N/A'),
            'Values Match': 'Yes' if plan1_param.get('compiled_value') == plan2_param.get('compiled_value') else 'No'
        })

    df = pd.DataFrame(rows)
    return df


def create_node_details_sheet(plan_data: dict, plan_name: str) -> pd.DataFrame:
    """Create node details sheet with detailed information for each operator node."""

    def remove_brackets(text):
        """Remove square brackets from text."""
        if not text:
            return ''
        # First remove empty brackets followed by period: [].
        result = str(text).replace('[].', '')
        # Then remove remaining [ and ]
        result = result.replace('[', '').replace(']', '')
        return result

    def extract_table_index_name(full_path):
        """Extract just the table/index name from a three-part naming convention.

        Examples:
            'QA07_Greg.dbo.AR_ChargeRecords.CIDX_AR_ChargeRecords_PatientID' -> 'CIDX_AR_ChargeRecords_PatientID'
            'dbo.TableName' -> 'TableName'
            'SimpleTable' -> 'SimpleTable'
        """
        if not full_path:
            return ''

        # Split by period to check for three-part naming
        parts = full_path.split('.')

        # If we have 3 or more parts (database.schema.table.index or database.schema.table)
        # Return the last part (table or index name)
        if len(parts) >= 3:
            return parts[-1]

        # Otherwise, return the original value
        return full_path

    rows = []

    for stmt in plan_data.get('statements', []):
        # Get node details if available
        node_details = stmt.get('node_details', [])

        for node in node_details:
            full_table_index = remove_brackets(node.get('table_index', ''))
            short_table_index = extract_table_index_name(full_table_index)

            row = {
                'Statement ID': node.get('statement_id', ''),
                'Node ID': node.get('node_id', ''),
                'Node Type': node.get('node_type', ''),
                'Table/Index Name': short_table_index,
                'Seek Predicates': remove_brackets(node.get('seek_predicates', '')),
                'Predicate': remove_brackets(node.get('predicate', '')),
                'Output List': remove_brackets(node.get('output_list', '')),
                'Physical Op': node.get('physical_op', ''),
                'Logical Op': node.get('logical_op', ''),
                'Est. Cost': node.get('estimated_cost', 0),
                'Est. CPU Cost': node.get('estimated_cpu_cost', 0),
                'Est. I/O Cost': node.get('estimated_io_cost', 0),
                'Est. Executions': node.get('estimated_executions', 0),
                'Est. Rows': node.get('estimated_rows', 0),
                'Actual Rows': node.get('actual_rows', 0),
                'Actual Executions': node.get('actual_executions', 0),
                'Actual Rebinds': node.get('actual_rebinds', 0),
                'Actual Rewinds': node.get('actual_rewinds', 0),
                'Parallel': node.get('parallel', 'No'),
                'Warnings': node.get('warnings', ''),
                'Table/Index Full Path': full_table_index
            }
            rows.append(row)

    if not rows:
        # Return empty DataFrame with correct columns
        return pd.DataFrame(columns=[
            'Statement ID', 'Node ID', 'Node Type', 'Table/Index Name',
            'Seek Predicates', 'Predicate', 'Output List', 'Physical Op', 'Logical Op',
            'Est. Cost', 'Est. CPU Cost', 'Est. I/O Cost', 'Est. Executions',
            'Est. Rows', 'Actual Rows', 'Actual Executions', 'Actual Rebinds', 'Actual Rewinds',
            'Parallel', 'Warnings', 'Table/Index Full Path'
        ])

    return pd.DataFrame(rows)


def format_excel_file(excel_file: str, config: ConfigLoader):
    """
    Apply formatting to the Excel file using openpyxl.

    Args:
        excel_file: Path to Excel file
        config: ConfigLoader instance
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(excel_file)

        # Define styles from config
        header_color = config.get_header_color()
        header_font_color = config.get_header_font_color()
        header_font_size = config.get_header_font_size()

        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        header_font = Font(bold=True, color=header_font_color, size=header_font_size)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Light red fill for cardinality estimation errors
        light_red_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')

        # Get thresholds from config (with defaults)
        cardinality_threshold = 200  # Could be made configurable
        cardinality_multiplier = 10  # Could be made configurable

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Color-code tabs: Light red for Statements, Light blue for Node Details
            if sheet_name.startswith('Stmts-'):
                ws.sheet_properties.tabColor = "FFC7CE"  # Light red
            elif sheet_name.startswith('Dtl-'):
                ws.sheet_properties.tabColor = "B4C7E7"  # Light blue

            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Apply borders to all cells and calculate column widths simultaneously
            column_widths = {}
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border

                    # Calculate max width for this column
                    col_letter = get_column_letter(cell.column)
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            column_widths[col_letter] = max(column_widths.get(col_letter, 0), cell_length)
                        except (AttributeError, TypeError) as e:
                            logging.debug(f"Could not calculate length for cell {cell.coordinate}: {e}")

            # Apply cardinality highlighting for Statements sheets
            if sheet_name.startswith('Stmts-'):
                apply_cardinality_highlighting(
                    ws, 'Estimated Rows', 'Actual Rows',
                    light_red_fill, cardinality_threshold, cardinality_multiplier
                )

            # Apply cardinality highlighting for Node Details sheets
            if sheet_name.startswith('Dtl-'):
                apply_cardinality_highlighting(
                    ws, 'Est. Rows', 'Actual Rows',
                    light_red_fill, cardinality_threshold, cardinality_multiplier
                )

            # Add autofilter to header row (only if there's data)
            if ws.max_row > 1 and ws.max_column > 0:
                ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

            # Freeze panes (header row)
            ws.freeze_panes = 'A2'

            # Set column widths based on calculated values
            min_col_width = 10  # Could be made configurable
            max_col_width = 100  # Could be made configurable

            for col_letter, max_length in column_widths.items():
                # Add extra space for filter dropdown arrow in header (3 extra chars)
                adjusted_width = max(min_col_width, min(max_length + 3, max_col_width))
                ws.column_dimensions[col_letter].width = adjusted_width
        
        wb.save(excel_file)
        logging.info(f"✅ Excel formatting applied successfully!")

    except ImportError:
        logging.warning("openpyxl not installed. Excel file created without formatting.")
        logging.info("   Install with: pip install openpyxl")
    except Exception as e:
        logging.warning(f"Could not apply formatting: {e}")


def main():
    """Main function to export JSON to Excel."""

    logging.info("="*80)
    logging.info("EXPORT TO EXCEL - STARTED")
    logging.info("="*80)
    logging.info(f"Log file: {LOG_FILE}")

    # Create output directory if it doesn't exist
    output_dir = config.get_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)

    # Define file paths
    json_filename = config.get_json_output_file()
    json_file = output_dir / json_filename

    # Check if JSON file exists
    if not json_file.exists():
        logging.error(f"{json_file} not found!")
        logging.info("   Please run 01_analyze_execution_plans.py first.")
        sys.exit(1)

    logging.info(f"Loading data from {json_file}...")
    try:
        data = load_json_data(str(json_file))
    except (FileNotFoundError, json.JSONDecodeError) as e:
        logging.error(f"Failed to load JSON file: {e}")
        sys.exit(1)

    # Validate JSON structure
    is_valid, error_msg = validate_json_structure(data)
    if not is_valid:
        logging.error(f"Invalid JSON structure: {error_msg}")
        sys.exit(1)

    # Determine which structure we're using
    try:
        plan1_key, plan2_key = get_plan_keys(data)

        plan1_data = safe_get_plan_data(data, plan1_key)
        plan2_data = safe_get_plan_data(data, plan2_key)

        if not plan1_data or not plan2_data:
            logging.error("Could not retrieve plan data from JSON")
            sys.exit(1)

        plan1_name = plan1_data.get('config_name', plan1_data.get('plan_name', 'Plan 1'))
        plan2_name = plan2_data.get('config_name', plan2_data.get('plan_name', 'Plan 2'))

    except ValueError as e:
        logging.error(f"Error determining plan structure: {e}")
        sys.exit(1)

    # Create Excel filename with plan names and timestamp
    # Sanitize plan names for filename
    safe_plan1_name = sanitize_filename(plan1_name)
    safe_plan2_name = sanitize_filename(plan2_name)

    timestamp_format = config.get_timestamp_format()
    timestamp = datetime.now().strftime(timestamp_format)

    # Parse timestamp
    date_part, time_part = parse_timestamp(timestamp)

    if time_part:
        excel_file = output_dir / f'Compare.{safe_plan1_name} vs {safe_plan2_name}.{date_part}.{time_part}.xlsx'
    else:
        excel_file = output_dir / f'Compare.{safe_plan1_name} vs {safe_plan2_name}.{date_part}.xlsx'

    # Create Excel writer
    excel_engine = config.get_excel_engine()
    max_sheet_name_length = config.get_max_sheet_name_length()

    # Track used sheet names to ensure uniqueness
    used_sheet_names = set()

    logging.info(f"Creating Excel file: {excel_file}...")
    with pd.ExcelWriter(str(excel_file), engine=excel_engine) as writer:

        # Sheet 1: Summary Comparison
        logging.info("  Creating Summary sheet...")
        df_summary = create_summary_sheet(data)
        sheet_name = create_unique_sheet_name('Summary', max_sheet_name_length, used_sheet_names)
        df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

        # Sheet 2: Plan 1 Statements
        logging.info(f"  Creating {plan1_name} Statements sheet...")
        df_plan1_stmts = create_statements_sheet(data[plan1_key], plan1_name)
        sheet_name_1 = create_unique_sheet_name(f'Stmts-{plan1_name}', max_sheet_name_length, used_sheet_names)
        df_plan1_stmts.to_excel(writer, sheet_name=sheet_name_1, index=False)

        # Sheet 3: Plan 2 Statements
        logging.info(f"  Creating {plan2_name} Statements sheet...")
        df_plan2_stmts = create_statements_sheet(data[plan2_key], plan2_name)
        sheet_name_2 = create_unique_sheet_name(f'Stmts-{plan2_name}', max_sheet_name_length, used_sheet_names)
        df_plan2_stmts.to_excel(writer, sheet_name=sheet_name_2, index=False)

        # Sheet 4: Plan 1 Node Details
        logging.info(f"  Creating {plan1_name} Node Details sheet...")
        df_plan1_nodes = create_node_details_sheet(data[plan1_key], plan1_name)
        sheet_name_1_nodes = create_unique_sheet_name(f'Dtl-{plan1_name}', max_sheet_name_length, used_sheet_names)
        df_plan1_nodes.to_excel(writer, sheet_name=sheet_name_1_nodes, index=False)

        # Sheet 5: Plan 2 Node Details
        logging.info(f"  Creating {plan2_name} Node Details sheet...")
        df_plan2_nodes = create_node_details_sheet(data[plan2_key], plan2_name)
        sheet_name_2_nodes = create_unique_sheet_name(f'Dtl-{plan2_name}', max_sheet_name_length, used_sheet_names)
        df_plan2_nodes.to_excel(writer, sheet_name=sheet_name_2_nodes, index=False)

        # Sheet 6: Missing Indexes
        logging.info("  Creating Missing Indexes sheet...")
        df_missing = create_missing_indexes_sheet(data)
        sheet_name = create_unique_sheet_name('Missing Indexes', max_sheet_name_length, used_sheet_names)
        df_missing.to_excel(writer, sheet_name=sheet_name, index=False)

        # Sheet 7: Warnings
        logging.info("  Creating Warnings sheet...")
        df_warnings = create_warnings_sheet(data)
        sheet_name = create_unique_sheet_name('Warnings', max_sheet_name_length, used_sheet_names)
        df_warnings.to_excel(writer, sheet_name=sheet_name, index=False)

        # Sheet 8: Parameters (if available)
        df_parameters = create_parameters_sheet(data)
        if not df_parameters.empty:
            logging.info("  Creating Parameters sheet...")
            sheet_name = create_unique_sheet_name('Parameters', max_sheet_name_length, used_sheet_names)
            df_parameters.to_excel(writer, sheet_name=sheet_name, index=False)

        # Sheet 9: Plan Overview (moved to last)
        logging.info("  Creating Plan Overview sheet...")
        df_overview = create_plan_overview_sheet(data)
        sheet_name = create_unique_sheet_name('Plan Overview', max_sheet_name_length, used_sheet_names)
        df_overview.to_excel(writer, sheet_name=sheet_name, index=False)

    logging.info(f"✅ Excel file created: {excel_file}")

    # Apply formatting
    logging.info("Applying formatting...")
    format_excel_file(str(excel_file), config)
    
    # Open the Excel file (non-blocking)
    open_excel_file(excel_file)

    logging.info("\n" + "="*80)
    logging.info("EXPORT COMPLETE")
    logging.info("="*80)
    logging.info(f"\nExcel file: {excel_file}")
    logging.info("\nSheets created:")
    logging.info("  1. Summary - Metric-by-metric comparison")
    logging.info(f"  2. Stmts-{plan1_name} - Detailed statement breakdown")
    logging.info(f"  3. Stmts-{plan2_name} - Detailed statement breakdown")
    logging.info(f"  4. Dtl-{plan1_name} - Operator-level details")
    logging.info(f"  5. Dtl-{plan2_name} - Operator-level details")
    logging.info("  6. Missing Indexes - Index recommendations")
    logging.info("  7. Warnings - Plan warnings and issues")
    logging.info("  8. Parameters - Compiled and runtime parameter values (if available)")
    logging.info("  9. Plan Overview - High-level summary of both plans")
    logging.info("="*80)


if __name__ == '__main__':
    main()

