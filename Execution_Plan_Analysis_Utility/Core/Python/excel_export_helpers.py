"""
Shared Excel Export Helper Module

This module provides common functionality for exporting execution plan data to Excel.
It is used by both the comparison export and single plan export scripts.

Author: SQL Server Toolkit
Date: 2026-03-20
"""

import logging
from typing import Dict, List, Any, Set, Tuple, Optional
from pathlib import Path


# ============================================================================
# CONSTANTS
# ============================================================================

# Cardinality estimation thresholds
CARDINALITY_THRESHOLD = 200  # Minimum row count to trigger highlighting
CARDINALITY_MULTIPLIER = 10  # Difference multiplier (10x)

# Column width constraints
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 100

# Timestamp format
TIMESTAMP_MIN_LENGTH = 15  # YYYYMMDD_HHMMSS


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def remove_brackets(text: str) -> str:
    """
    Remove square brackets from text.

    Args:
        text: Text potentially containing brackets

    Returns:
        Text with brackets removed
    """
    if not text:
        return ''
    # First remove empty brackets followed by period: [].
    result = str(text).replace('[].', '')
    # Then remove remaining [ and ]
    result = result.replace('[', '').replace(']', '')
    return result


def extract_table_index_name(full_path: str) -> str:
    """
    Extract just the table/index name from a three-part naming convention.

    Examples:
        'QA07_Greg.dbo.AR_ChargeRecords.CIDX_AR_ChargeRecords_PatientID'
            -> 'CIDX_AR_ChargeRecords_PatientID'
        'dbo.TableName' -> 'TableName'
        'SimpleTable' -> 'SimpleTable'

    Args:
        full_path: Full database path

    Returns:
        Extracted table/index name
    """
    if not full_path:
        return ''

    # Split by period to check for three-part naming
    parts = full_path.split('.')

    # If we have 3 or more parts (database.schema.table.index or database.schema.table)
    # Return the last part (table or index name)
    if len(parts) >= 3:
        return parts[-1]

    # Otherwise, return the original value
    return full_path


def sanitize_filename(name: str) -> str:
    """
    Sanitize a string for use in filenames.

    Args:
        name: Original name

    Returns:
        Sanitized name safe for filenames (returns 'unnamed' if result is empty)
    """
    if not name:
        return 'unnamed'

    result = "".join(c for c in name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')

    # Return default if result is empty (all special characters)
    return result if result else 'unnamed'


def create_unique_sheet_name(base_name: str, max_length: int, existing_names: Set[str]) -> str:
    """
    Create unique sheet name with truncation.

    Args:
        base_name: Desired sheet name
        max_length: Maximum allowed length
        existing_names: Set of already used names

    Returns:
        Unique sheet name within length limit
    """
    name = base_name[:max_length]
    counter = 1
    original_name = name

    while name in existing_names:
        suffix = f"_{counter}"
        name = original_name[:max_length - len(suffix)] + suffix
        counter += 1

    existing_names.add(name)
    return name


def parse_timestamp(timestamp: str) -> Tuple[str, str]:
    """
    Parse timestamp into date and time parts.

    Args:
        timestamp: Timestamp string (expected format: YYYYMMDD_HHMMSS)

    Returns:
        Tuple of (date_part, time_part)
    """
    try:
        # Assume format is YYYYMMDD_HHMMSS
        if '_' in timestamp and len(timestamp) >= TIMESTAMP_MIN_LENGTH:
            date_part = timestamp[:8]  # YYYYMMDD
            time_part = timestamp[9:]  # HHMMSS
            return date_part, time_part
        else:
            # Fallback if format is different
            return timestamp, ''
    except Exception as e:
        logging.warning(f"Could not parse timestamp '{timestamp}': {e}")
        return timestamp, ''


def apply_cardinality_highlighting(ws: Any, est_col_name: str, act_col_name: str,
                                   light_red_fill: Any, threshold: int = CARDINALITY_THRESHOLD,
                                   multiplier: int = CARDINALITY_MULTIPLIER) -> None:
    """
    Apply cardinality estimation error highlighting to a worksheet.

    Highlights cells where estimated vs actual rows differ by the multiplier (default 10x)
    and at least one value exceeds the threshold (default 200).

    Args:
        ws: Worksheet object (openpyxl.worksheet.worksheet.Worksheet)
        est_col_name: Name of estimated rows column
        act_col_name: Name of actual rows column
        light_red_fill: Fill style for highlighting (openpyxl.styles.PatternFill)
        threshold: Minimum row count to trigger highlighting (default 200)
        multiplier: Difference multiplier to trigger highlighting (default 10)
    """
    try:
        header_row = [cell.value for cell in ws[1]]
        est_rows_col = header_row.index(est_col_name) + 1
        act_rows_col = header_row.index(act_col_name) + 1

        # Iterate through data rows (skip header)
        for row_idx in range(2, ws.max_row + 1):
            est_rows_cell = ws.cell(row=row_idx, column=est_rows_col)
            act_rows_cell = ws.cell(row=row_idx, column=act_rows_col)

            est_rows = est_rows_cell.value
            act_rows = act_rows_cell.value

            # Convert to numbers, handle None/empty values
            try:
                est_rows = float(est_rows) if est_rows is not None else 0
                act_rows = float(act_rows) if act_rows is not None else 0
            except (ValueError, TypeError):
                logging.debug(f"Could not convert row {row_idx} values to float: est={est_rows}, act={act_rows}")
                continue

            # Check for multiplier difference
            has_significant_difference = False
            if est_rows > 0 and act_rows >= multiplier * est_rows:
                has_significant_difference = True  # Underestimation
            elif act_rows > 0 and est_rows >= multiplier * act_rows:
                has_significant_difference = True  # Overestimation
            elif est_rows > 0 and act_rows == 0:
                has_significant_difference = True  # Overestimation (got nothing)
            elif act_rows > 0 and est_rows == 0:
                has_significant_difference = True  # Underestimation (expected nothing)

            # Only highlight if there's a significant difference AND at least one value > threshold
            if has_significant_difference and (est_rows > threshold or act_rows > threshold):
                est_rows_cell.fill = light_red_fill
                act_rows_cell.fill = light_red_fill

    except ValueError as e:
        # Columns not found in this sheet, skip highlighting
        logging.debug(f"Columns '{est_col_name}' or '{act_col_name}' not found in sheet: {e}")



def format_worksheet(ws: Any, header_fill: Any, header_font: Any, border: Any, config: Any,
                     light_red_fill: Optional[Any] = None, apply_highlighting: bool = False,
                     est_col_name: str = 'Estimated Rows', act_col_name: str = 'Actual Rows') -> None:
    """
    Apply standard formatting to a worksheet.

    Args:
        ws: Worksheet object (openpyxl.worksheet.worksheet.Worksheet)
        header_fill: Fill style for header row (openpyxl.styles.PatternFill)
        header_font: Font style for header row (openpyxl.styles.Font)
        border: Border style for all cells (openpyxl.styles.Border)
        config: ConfigLoader instance
        light_red_fill: Optional fill style for cardinality highlighting (openpyxl.styles.PatternFill)
        apply_highlighting: Whether to apply cardinality highlighting
        est_col_name: Name of estimated rows column (for highlighting)
        act_col_name: Name of actual rows column (for highlighting)
    """
    try:
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Apply borders to all cells and calculate column widths simultaneously
        column_widths = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

                # Calculate max width for this column
                col_letter = get_column_letter(cell.column)
                if cell.value:
                    try:
                        cell_length = len(str(cell.value))
                        column_widths[col_letter] = max(column_widths.get(col_letter, 0), cell_length)
                    except (AttributeError, TypeError) as e:
                        logging.debug(f"Could not calculate length for cell {cell.coordinate}: {e}")

        # Apply cardinality highlighting if requested
        if apply_highlighting and light_red_fill:
            apply_cardinality_highlighting(ws, est_col_name, act_col_name, light_red_fill)

        # Add autofilter to header row (only if there's data)
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'

        # Freeze panes (header row)
        ws.freeze_panes = 'A2'

        # Set column widths based on calculated values
        for col_letter, max_length in column_widths.items():
            # Add extra space for filter dropdown arrow in header (3 extra chars)
            adjusted_width = max(MIN_COLUMN_WIDTH, min(max_length + 3, MAX_COLUMN_WIDTH))
            ws.column_dimensions[col_letter].width = adjusted_width

    except Exception as e:
        logging.error(f"Error formatting worksheet: {e}")


def open_excel_file(excel_file: Path):
    """
    Open an Excel file in the default application.

    Args:
        excel_file: Path to Excel file
    """
    import sys
    import os

    logging.info(f"\nOpening {excel_file}...")
    try:
        import subprocess

        if sys.platform == 'win32':
            # Use os.startfile for Windows (safer than shell=True)
            try:
                os.startfile(str(excel_file))
            except AttributeError:
                # Fallback if os.startfile not available (shouldn't happen on Windows)
                subprocess.Popen(['cmd', '/c', 'start', '', str(excel_file)])
        elif sys.platform == 'darwin':  # macOS
            subprocess.Popen(['open', str(excel_file)])
        else:  # linux
            subprocess.Popen(['xdg-open', str(excel_file)])

        logging.info("✅ Excel file opened successfully!")
    except Exception as e:
        logging.warning(f"Could not automatically open Excel file: {e}")
        logging.info(f"   Please open {excel_file} manually.")


def set_tab_colors(wb: Any, config: Any, plan1_name: Optional[str] = None,
                   plan2_name: Optional[str] = None) -> None:
    """
    Set tab colors for worksheets based on naming conventions.

    Args:
        wb: Workbook object (openpyxl.workbook.workbook.Workbook)
        config: ConfigLoader instance
        plan1_name: Optional name of first plan (for matching sheet names)
        plan2_name: Optional name of second plan (for matching sheet names)
    """
    try:
        summary_color = config.get_summary_tab_color()
        plan1_color = config.get_plan1_tab_color()
        plan2_color = config.get_plan2_tab_color()

        for sheet in wb.worksheets:
            sheet_name = sheet.title

            # Summary and overview sheets
            if sheet_name in ['Summary', 'Plan Overview', 'Weighted Score', 'Missing Indexes', 'Warnings']:
                sheet.sheet_properties.tabColor = summary_color
            # Try to match plan names in sheet names
            elif plan1_name and plan1_name in sheet_name:
                sheet.sheet_properties.tabColor = plan1_color
            elif plan2_name and plan2_name in sheet_name:
                sheet.sheet_properties.tabColor = plan2_color
            # Fallback: Use position-based coloring for Stmts-* and Dtl-* sheets
            elif sheet_name.startswith('Stmts-') or sheet_name.startswith('Dtl-'):
                # Alternate colors based on position
                if wb.worksheets.index(sheet) % 2 == 0:
                    sheet.sheet_properties.tabColor = plan1_color
                else:
                    sheet.sheet_properties.tabColor = plan2_color
            # Default to summary color for any other sheets
            else:
                sheet.sheet_properties.tabColor = summary_color

    except Exception as e:
        logging.debug(f"Could not set tab colors: {e}")

