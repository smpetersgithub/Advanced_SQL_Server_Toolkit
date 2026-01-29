import os
import re
import csv
from collections import defaultdict
import configparser
import logging
from datetime import datetime

# Patterns
sp_pattern1 = re.compile(r'prepareCall\s*\(\s*["\{]+\s*(?:\?=)?call\s+(?:dbo\.)?([a-zA-Z0-9_]+)', re.IGNORECASE)
sp_pattern2 = re.compile(r'getCallableStatement\s*\(\s*["\{]+\s*(?:\?=)?call\s+(?:dbo\.)?([a-zA-Z0-9_]+)', re.IGNORECASE)
sp_pattern3 = re.compile(r'execute\s*\(\s*["\']([a-zA-Z0-9_]+)["\']', re.IGNORECASE)
sp_pattern4 = re.compile(r'super\s*\(\s*ds\s*,\s*["\']([a-zA-Z0-9_]+)["\']', re.IGNORECASE)
sp_pattern5 = re.compile(r'SQL\s*=\s*["\'](?:dbo\.)?([a-zA-Z0-9_]+)["\']', re.IGNORECASE)

class_pattern = re.compile(r'public\s+class\s+([a-zA-Z0-9_]+)')
import_pattern = re.compile(r'import\s+[\w.]+\.([a-zA-Z0-9_]+DAO);')
new_dao_pattern = re.compile(r'new\s+([a-zA-Z0-9_]+DAO)\s*\(')

# Data structures
dao_to_stored_procs = defaultdict(set)  # DAO -> Set of stored procs
dao_to_file = {}  # DAO -> file path
ui_to_daos = defaultdict(set)  # UI component -> Set of DAOs used
ui_to_file = {}  # UI component -> file path
ui_types = {}  # UI component -> type (Handler, Action, etc.)

def scan_dao_file(filepath, dao_count_tracker):
    """Scan a DAO file for stored procedures"""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            
        class_match = class_pattern.search(content)
        if not class_match:
            return
            
        class_name = class_match.group(1)
        
        # Only process DAO classes
        if not class_name.endswith('DAO'):
            return 0

        dao_count_tracker[0] += 1

        # Find all stored procedures
        sp_matches = []
        sp_matches.extend(sp_pattern1.findall(content))
        sp_matches.extend(sp_pattern2.findall(content))
        
        for match in sp_pattern3.findall(content):
            if match.lower().startswith('sp'):
                sp_matches.append(match)
        for match in sp_pattern4.findall(content):
            if match.lower().startswith('sp'):
                sp_matches.append(match)
        for match in sp_pattern5.findall(content):
            if match.lower().startswith('sp'):
                sp_matches.append(match)
        
        if sp_matches:
            dao_to_file[class_name] = filepath
            dao_to_stored_procs[class_name].update(sp_matches)
            return 1
        return 0

    except Exception as e:
        return 0

def scan_ui_file(filepath):
    """Scan UI files (Handlers, Actions, etc.) for DAO usage"""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            
        class_match = class_pattern.search(content)
        if not class_match:
            return
            
        class_name = class_match.group(1)
        
        # Determine UI type
        ui_type = 'Unknown'
        if 'Handler' in class_name:
            ui_type = 'Page Handler'
        elif 'Action' in class_name:
            ui_type = 'Struts Action'
        elif 'Controller' in class_name:
            ui_type = 'Spring Controller'
        elif 'Service' in class_name or 'ServiceImpl' in class_name:
            ui_type = 'Service'
        else:
            return  # Skip non-UI files
            
        # Find DAO imports
        imported_daos = set(import_pattern.findall(content))
        
        # Find DAO instantiations (new SomeDAO())
        instantiated_daos = set(new_dao_pattern.findall(content))
        
        # Combine both
        used_daos = imported_daos | instantiated_daos
        
        if used_daos:
            ui_to_file[class_name] = filepath
            ui_types[class_name] = ui_type
            ui_to_daos[class_name].update(used_daos)

    except Exception as e:
        return 0

def main():
    # Load configuration
    config = configparser.ConfigParser()
    config_path = os.path.join(os.path.dirname(__file__), 'config.ini')
    config.read(config_path)

    # Get configuration values
    java_source_dirs = [
        config.get('Paths', 'java_source_dir_1'),
        config.get('Paths', 'java_source_dir_2')
    ]
    output_dir = config.get('Paths', 'output_dir')
    csv_filename = config.get('Files', 'complete_mapping_csv')
    excel_filename = config.get('Files', 'complete_mapping_excel')
    log_dir = config.get('Paths', 'log_dir')

    # Setup logging
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    log_filename = f'log_01_extract_complete_ui_mapping_{timestamp}.log'
    log_filepath = os.path.join(log_dir, log_filename)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filepath, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    logging.info("=" * 80)
    logging.info("Stored Procedure Java Tracer - Complete UI Mapping Generator")
    logging.info("Stored Procedure -> DAO -> UI Handler/Action -> JSP")
    logging.info("=" * 80)

    logging.info("\n[1/4] Scanning DAO files for stored procedures...")
    dao_count_tracker = [0]

    # Scan DAO files
    for base_dir in java_source_dirs:
        if os.path.exists(base_dir):
            for root, dirs, files in os.walk(base_dir):
                for file in files:
                    if file.endswith('DAO.java'):
                        scan_dao_file(os.path.join(root, file), dao_count_tracker)

    dao_count = dao_count_tracker[0]
    logging.info(f"   Found {dao_count} DAO files")
    logging.info(f"   Found {len(dao_to_stored_procs)} DAOs with stored procedures")

    logging.info("\n[2/4] Scanning UI files (Handlers, Actions, Services) for DAO usage...")
    ui_count = 0

    for base_dir in java_source_dirs:
        if os.path.exists(base_dir):
            for root, dirs, files in os.walk(base_dir):
                for file in files:
                    if file.endswith('.java') and not file.endswith('DAO.java'):
                        scan_ui_file(os.path.join(root, file))
                        if len(ui_to_daos) > ui_count:
                            ui_count = len(ui_to_daos)

    logging.info(f"   Found {ui_count} UI components using DAOs")

    logging.info("\n[3/4] Building complete mapping...")

    # Build the complete mapping: StoredProc -> DAO -> UI
    complete_mapping = []

    for dao_name, stored_procs in dao_to_stored_procs.items():
        dao_file = dao_to_file.get(dao_name, 'Unknown')

        # Find all UI components that use this DAO
        ui_components_using_dao = []
        for ui_name, daos_used in ui_to_daos.items():
            if dao_name in daos_used:
                ui_components_using_dao.append({
                    'ui_name': ui_name,
                    'ui_type': ui_types.get(ui_name, 'Unknown'),
                    'controller_file': ui_to_file.get(ui_name, 'Unknown')
                })

        # Create mapping entries
        for sp_name in stored_procs:
            if ui_components_using_dao:
                # DAO is used by UI components
                for ui_info in ui_components_using_dao:
                    complete_mapping.append({
                        'Stored_Procedure': sp_name,
                        'DAO_Class': dao_name,
                        'DAO_File': dao_file,
                        'UI_Component': ui_info['ui_name'],
                        'UI_Type': ui_info['ui_type'],
                        'Controller_File': ui_info['controller_file']
                    })
            else:
                # DAO not used by any UI (or we couldn't find it)
                complete_mapping.append({
                    'Stored_Procedure': sp_name,
                    'DAO_Class': dao_name,
                    'DAO_File': dao_file,
                    'UI_Component': 'Not Found',
                    'UI_Type': 'N/A',
                    'Controller_File': 'N/A'
                })

    logging.info(f"   Created {len(complete_mapping)} complete mappings")

    logging.info("\n[4/4] Writing output files...")

    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Write CSV - Complete mapping
    csv_file = os.path.join(output_dir, csv_filename)
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        fieldnames = ['Stored_Procedure', 'DAO_Class', 'DAO_File', 'UI_Component', 'UI_Type', 'Controller_File']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in sorted(complete_mapping, key=lambda x: (x['Stored_Procedure'], x['DAO_Class'], x['UI_Component'])):
            writer.writerow(row)

    logging.info(f"   CSV: {csv_file}")

    # Write Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "StoredProc to UI"

        headers = ['Stored Procedure', 'DAO Class', 'DAO File', 'UI Component', 'UI Type', 'UI File']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row in sorted(complete_mapping, key=lambda x: (x['Stored_Procedure'], x['DAO_Class'], x['UI_Component'])):
            ws.append([row['Stored_Procedure'], row['DAO_Class'], row['DAO_File'],
                       row['UI_Component'], row['UI_Type'], row['Controller_File']])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width

        excel_file = os.path.join(output_dir, excel_filename)
        wb.save(excel_file)
        logging.info(f"   Excel: {excel_file}")
    except ImportError:
        logging.info("   Excel: Skipped (openpyxl not available)")

    logging.info("\n" + "=" * 80)
    logging.info("SUMMARY")
    logging.info("=" * 80)
    logging.info(f"DAOs scanned: {dao_count}")
    logging.info(f"DAOs with stored procs: {len(dao_to_stored_procs)}")
    logging.info(f"UI components found: {ui_count}")
    logging.info(f"Total mappings: {len(complete_mapping)}")
    logging.info(f"Unique stored procedures: {len(set(m['Stored_Procedure'] for m in complete_mapping))}")
    logging.info(f"Log file: {log_filepath}")
    logging.info("\nDone!")

if __name__ == "__main__":
    main()


